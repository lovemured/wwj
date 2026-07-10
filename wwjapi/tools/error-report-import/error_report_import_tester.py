#!/usr/bin/env python3
"""CRM Excel import helper for Lead/error-report import testing."""

import argparse
import json
import mimetypes
import os
import re
import threading
import time
import uuid
from datetime import datetime, timedelta
from pathlib import Path

import requests
from openpyxl import load_workbook


DEFAULT_OUT_DIR = Path("outputs/error-report-import")
DEFAULT_FAYE_URL = "https://faye-dev.ikcrm.com/faye"
MODULE_CONFIGS = {
    "lead": {"loader_name": "Lead", "api_path": "leads", "label": "线索"},
    "customer": {"loader_name": "Customer", "api_path": "customers", "label": "客户"},
    "contact": {"loader_name": "Contact", "api_path": "contacts", "label": "联系人"},
    "opportunity": {"loader_name": "Opportunity", "api_path": "opportunities", "label": "商机"},
    "quotation": {"loader_name": "Quotation", "api_path": "quotations", "label": "报价单"},
    "contract": {"loader_name": "Contract", "api_path": "contracts", "label": "合同"},
    "payment-plan": {"loader_name": "ReceivedPaymentPlan", "api_path": "received_payment_plans", "label": "回款计划"},
    "received-payment": {"loader_name": "ReceivedPayment", "api_path": "received_payments", "label": "回款记录"},
    "invoiced-payment": {"loader_name": "InvoicedPayment", "api_path": "invoiced_payments", "label": "开票记录"},
    "product": {"loader_name": "Product", "api_path": "products", "label": "产品"},
}
MODULE_ALIASES = {
    "线索": "lead",
    "客户": "customer",
    "联系人": "contact",
    "商机": "opportunity",
    "报价单": "quotation",
    "合同": "contract",
    "回款计划": "payment-plan",
    "回款记录": "received-payment",
    "开票记录": "invoiced-payment",
    "产品": "product",
}
BUSINESS_TYPE_MODULES = {"customer", "opportunity", "quotation", "contract"}
ENV_CONFIGS = {
    "test": {
        "api": "https://lxcrm-test.weiwenjia.com",
        "faye_url": DEFAULT_FAYE_URL,
        "entity_loader_prefix": "/api/pc/entity_loaders",
        "entity_loader_new_prefix": "/api/pc/entity_loaders",
        "entity_loader_history_prefix": "/api/pc/entity_loaders",
        "qiniu_token_path": "/api/pc/qiniu/auth/oss_upload_token.json",
    },
    "staging": {
        "api": "https://lxcrm-staging.weiwenjia.com",
        "faye_url": DEFAULT_FAYE_URL,
        "entity_loader_prefix": "/api/entity_loaders",
        "entity_loader_new_prefix": "/api/pc/entity_loaders",
        "entity_loader_history_prefix": "/api/pc/entity_loaders",
        "qiniu_token_path": "/api/qiniu/auth/oss_upload_token.json",
    },
}


def pc_url(api):
    return api.rstrip("/").replace("//lxcrm-test.", "//lxcrm-api-test.").replace(
        "//lxcrm-staging.", "//lxcrm-api-staging."
    )


def auth_headers(token):
    return {
        "Authorization": f"Token token={token}",
        "ACCESS-TOKEN": token,
        "User-Agent": "wwjapi-import-test/1.0",
    }


def print_json(title, data):
    print(f"\n## {title}")
    print(json.dumps(data, ensure_ascii=False, indent=2))


def normalize_module(module):
    if not module:
        return "lead"
    key = MODULE_ALIASES.get(module, module).strip().lower().replace("_", "-")
    if key not in MODULE_CONFIGS:
        choices = ", ".join(MODULE_CONFIGS)
        raise SystemExit(f"未知模块: {module}，可选: {choices}")
    return key


def request_json(method, url, *, token, **kwargs):
    headers = kwargs.pop("headers", {})
    merged = auth_headers(token)
    merged.update(headers)
    resp = requests.request(method, url, headers=merged, timeout=kwargs.pop("timeout", 60), **kwargs)
    try:
        data = resp.json()
    except ValueError:
        data = {"raw_text": resp.text[:2000]}
    return resp.status_code, data


def parse_business_type_map(value):
    result = {}
    if not value:
        return result
    for item in value.split(","):
        if not item.strip():
            continue
        if "=" not in item:
            raise SystemExit(f"业务类型映射格式错误: {item}，应为 module=id 或 module=name")
        module, template = item.split("=", 1)
        result[normalize_module(module.strip())] = template.strip()
    return result


def business_type_value(args):
    mapped = parse_business_type_map(getattr(args, "business_type_map", None)).get(args.module_key)
    return mapped or getattr(args, "business_type_id", None) or getattr(args, "business_type_name", None)


def list_business_types(args):
    model_klass = args.module_config["loader_name"]
    status, data = request_json(
        "GET",
        f"{pc_url(args.api)}/api/pc/custom_fields",
        token=args.token,
        params={"model_klass": model_klass},
        timeout=30,
    )
    groups = (data.get("data") or {}).get("custom_field_groups") or []
    field_id = None
    for group in groups:
        for field in group.get("custom_fields") or []:
            if field.get("field_id"):
                field_id = field["field_id"]
                break
        if field_id:
            break
    if status != 200 or not field_id:
        return []

    status, detail = request_json(
        "GET",
        f"{pc_url(args.api)}/api/pc/custom_fields/{field_id}",
        token=args.token,
        params={"custom_field_template_id": 1331},
        timeout=30,
    )
    if status != 200:
        return []
    return (detail.get("data") or {}).get("custom_field_templates") or []


def resolve_business_type(args):
    if getattr(args, "business_type_resolved", False):
        return
    if args.module_key not in BUSINESS_TYPE_MODULES:
        args.custom_field_template_id = ""
        args.business_type_resolved = True
        return

    requested = business_type_value(args)
    templates = list_business_types(args)
    enabled = [item for item in templates if item.get("status") == "enable"]
    selected = None
    if requested:
        for item in enabled:
            if str(item.get("id")) == str(requested) or item.get("name") == requested:
                selected = item
                break
        if not selected:
            names = ", ".join(f"{item.get('id')}:{item.get('name')}" for item in enabled)
            raise RuntimeError(f"{args.module_config['label']} 未找到业务类型 {requested}，可选: {names}")
    elif enabled:
        selected = enabled[0]

    if not selected:
        args.custom_field_template_id = ""
        args.business_type_resolved = True
        print(f"{args.module_config['label']} 未发现启用业务类型，继续使用默认模板")
        return

    args.custom_field_template_id = str(selected["id"])
    args.business_type_resolved = True
    print(f"业务类型: {args.module_config['label']} -> {selected.get('name')}({selected.get('id')})")


def print_business_types(args):
    if args.module_key not in BUSINESS_TYPE_MODULES:
        print_json("business_types", {"module": args.module_key, "items": []})
        return []
    items = [
        {"id": item.get("id"), "name": item.get("name"), "status": item.get("status")}
        for item in list_business_types(args)
    ]
    print_json("business_types", {"module": args.module_key, "items": items})
    return items


def entity_loader_url(args, path):
    clean_path = path.lstrip("/")
    if clean_path == "new":
        prefix = args.entity_loader_new_prefix
    elif clean_path == "import_histories":
        prefix = args.entity_loader_history_prefix
    else:
        prefix = args.entity_loader_prefix
    return f"{args.api.rstrip('/')}{prefix}/{path.lstrip('/')}"


def current_user(api, token):
    status, data = request_json("GET", f"{pc_url(api)}/api/v2/user/info", token=token, timeout=30)
    if status != 200 or data.get("code") != 0:
        raise RuntimeError(f"获取当前用户失败: HTTP {status} {data}")
    info = data["data"]
    return str(info["id"]), str(info["organization_id"]), info


def download_template(args):
    resolve_business_type(args)
    out = Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)
    url = entity_loader_url(args, f"template/{args.loader_name}.xlsx")
    params = {}
    if getattr(args, "custom_field_template_id", ""):
        params["custom_field_template_id"] = args.custom_field_template_id
    resp = requests.get(url, headers=auth_headers(args.token), params=params, timeout=60)
    resp.raise_for_status()
    out.write_bytes(resp.content)
    print(f"模板已下载: {out} ({len(resp.content)} bytes)")


def find_template_rows(ws):
    marker_row = None
    for row in range(1, min(ws.max_row, 80) + 1):
        first = ws.cell(row, 1).value
        if isinstance(first, str) and "请从下面表格开始填写" in first:
            marker_row = row
            break
    if marker_row:
        return marker_row + 1, marker_row + 2

    # Fallback for older templates.
    for row in range(1, min(ws.max_row, 80) + 1):
        values = [ws.cell(row, col).value for col in range(1, min(ws.max_column, 30) + 1)]
        if any(isinstance(v, str) and "必填" in v for v in values):
            return row, row + 1
    raise RuntimeError("无法识别模板表头行")


def clean_header(header):
    return str(header or "").replace("(必填)", "").replace("（必填）", "").strip()


def first_option_from_sheet(wb, header):
    values = option_values_from_sheet(wb, header, limit=1)
    return values[0] if values else None


def option_values_from_sheet(wb, header, limit=2, example=None):
    name = clean_header(header)
    candidates = [f"Sheet{name}", f"Sheet_{name}"]
    if isinstance(example, str):
        candidates.extend(re.findall(r"Sheet[_\w\u4e00-\u9fff]+", example))
    normalized = [c.replace("_", "") for c in candidates]
    for ws in wb.worksheets:
        if ws.title in candidates or ws.title.replace("_", "") in normalized:
            values = []
            for row in range(2, ws.max_row + 1):
                value = ws.cell(row, 1).value
                if value not in (None, ""):
                    values.append(value)
                if len(values) >= limit:
                    break
            return values
    return []


def user_option_values(wb, limit=2):
    for sheet_name in ("Sheet负责人", "Sheet自定义用户"):
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            values = []
            for row in range(2, ws.max_row + 1):
                value = ws.cell(row, 1).value
                if value not in (None, ""):
                    values.append(value)
                if len(values) >= limit:
                    return values
    return []


def template_example_value(ws, col):
    value = ws.cell(8, col).value
    if isinstance(value, str) and value.startswith("("):
        return None
    return value if value not in (None, "") else None


def full_field_value(wb, ws, header, col, suffix, index, context=None, module_key=None):
    name = clean_header(header)
    context = context or {}
    if name == "上级客户":
        return None
    if name in ("对应客户", "客户名称") or name.startswith("对应客户"):
        return (context.get("customer") or {}).get("name") or f"自动化客户{suffix}-{index}"
    if name == "[对应客户]ID":
        return (context.get("customer") or {}).get("id")
    if name in ("对应商机", "商机标题") or name.startswith("对应商机"):
        return (context.get("opportunity") or {}).get("name") or f"自动化商机{suffix}-{index}"
    if name == "[对应商机]ID":
        return (context.get("opportunity") or {}).get("id")
    if name.startswith("合同标题") and context.get("contract") and module_key != "contract":
        return context["contract"].get("name")
    if name in ("[对应合同]ID", "[合同标题]ID"):
        return (context.get("contract") or {}).get("id")
    if name == "对应联系人":
        return (context.get("contact") or {}).get("name")
    if name == "[关联产品]ID":
        return (context.get("product") or {}).get("id")
    if name == "[关联产品]产品编号":
        return (context.get("product") or {}).get("no")
    if name == "[关联产品]产品名称":
        return (context.get("product") or {}).get("name")
    if "[关联产品]售价" in name:
        return 100
    if "[关联产品]数量" in name:
        return 1
    if "[关联产品]计算数字" in name:
        return 1
    if "[关联产品]单选计算" in name:
        option = first_option_from_sheet(wb, header)
        return option
    if "[关联产品]备注" in name:
        return f"关联产品备注{suffix}-{index}"
    if name in ("负责人", "经手人") or "用户单选" in name:
        users = user_option_values(wb, limit=1)
        if users:
            return users[0]
    if name == "协作人" or "用户多选" in name:
        users = user_option_values(wb, limit=2)
        if users:
            return "，".join(str(v) for v in users)

    options = option_values_from_sheet(wb, header, limit=2, example=ws.cell(8, col).value)
    if options:
        if "多选" in name or "用户多选" in name:
            return "，".join(str(v) for v in options)
        return options[0]

    if name == "ID" or "唯一性ID" in name:
        return None
    if "客户名称" in name:
        return f"自动化客户{suffix}-{index}"
    if "产品名称" in name:
        return f"自动化产品{suffix}-{index}"
    if "产品编号" in name:
        return f"PRD{suffix}{index:03d}"
    if "商机标题" in name or name == "商机名称":
        return f"自动化商机{suffix}-{index}"
    if "报价单名称" in name or "报价单主标题" in name:
        return f"自动化报价单{suffix}-{index}"
    if "报价单编号" in name:
        return f"QT{suffix}{index:03d}"
    if "合同标题" in name:
        return f"自动化合同{suffix}-{index}"
    if "合同编号" in name:
        return f"CT{suffix}{index:03d}"
    if "发票号码" in name:
        return f"INV{suffix}{index:03d}"
    if "回款期次" in name:
        return index
    if "产品费用合计" in name:
        return 100
    if "整单折扣" in name:
        return 1
    if "销售单位" in name or name == "单位":
        return "个"
    if "规格" in name:
        return f"规格{index}"
    if "签约人" in name:
        return f"签约人{index}"
    if "姓名" in name:
        return f"自动化姓名{suffix}-{index}"
    if "公司" in name:
        return f"自动化测试公司{suffix}-{index}"
    if name == "电话" or "电话" in name:
        return f"021-6{suffix[-5:]}{index}"
    if "手机" in name:
        return f"1390000{index:04d}"
    if "邮箱" in name:
        return f"lead{suffix}{index}@example.com"
    if "微信" in name:
        return f"wx{suffix}{index}"
    if "QQ" in name or "qq" in name:
        return f"88{suffix[-4:]}{index:02d}"
    if "旺旺" in name:
        return f"ww{suffix}{index}"
    if name == "国家":
        return "中国"
    if name == "省":
        return "上海"
    if name == "市":
        return "上海市"
    if name == "区":
        return "徐汇区"
    if "地址" in name:
        return f"上海市徐汇区自动化测试路{index}号"
    if "邮编" in name:
        return "200030"
    if "网址" in name or "链接" in name:
        return f"https://example.com/lead/{suffix}/{index}"
    if "日期" in name:
        return (datetime.now() + timedelta(days=index)).strftime("%Y-%m-%d")
    if "时间" in name:
        return (datetime.now() + timedelta(days=index)).strftime("%Y-%m-%d %H:%M")
    if "整数" in name:
        return index
    if "小数" in name:
        return round(index + 0.25, 2)
    if "金额" in name:
        return round(1000 + index * 10.5, 2)
    if "计算" in name:
        return 1
    if "百分比" in name:
        return 10
    if "备注" in name:
        return f"自动化导入备注{suffix}-{index}"
    if "多行" in name:
        return f"自动化多行文本{suffix}-{index}"
    if "单行" in name or "文本" in name:
        return f"自动化文本{suffix}-{index}"

    example = template_example_value(ws, col)
    return example if example is not None else f"自动化字段{suffix}-{index}"


def minimal_field_value(header, suffix, index):
    if "客户名称" in header:
        return f"自动化客户{suffix}-{index}"
    if "产品名称" in header:
        return f"自动化产品{suffix}-{index}"
    if "商机标题" in header or "商机名称" in header:
        return f"自动化商机{suffix}-{index}"
    if "报价单名称" in header or "报价单主标题" in header:
        return f"自动化报价单{suffix}-{index}"
    if "合同标题" in header:
        return f"自动化合同{suffix}-{index}"
    if "公司" in header:
        return f"自动化测试公司{suffix}-{index}"
    if "姓名" in header:
        return f"自动化姓名{suffix}-{index}"
    if "手机" in header and "自定义" not in header:
        return f"1390000{index:04d}"
    if "邮箱" in header:
        return f"lead{suffix}{index}@example.com"
    return None


def generate_source_from_template(args):
    src = Path(args.template)
    out = Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(src)
    ws = wb["线索XS"] if "线索XS" in wb.sheetnames else wb[wb.sheetnames[0]]
    header_row, data_start = find_template_rows(ws)
    if getattr(args, "module_key", "") == "customer":
        for col in range(1, ws.max_column + 1):
            ws.cell(data_start, col).value = ws.cell(header_row, col).value
        header_row = data_start
        data_start = header_row + 1
    headers = {str(ws.cell(header_row, col).value or ""): col for col in range(1, ws.max_column + 1)}

    for row in range(data_start, data_start + max(args.rows, 20)):
        for col in range(1, ws.max_column + 1):
            ws.cell(row, col).value = None

    suffix = time.strftime("%H%M%S")
    for i in range(1, args.rows + 1):
        row = data_start + i - 1
        for header, col in headers.items():
            if args.fill_mode == "all":
                value = full_field_value(
                    wb,
                    ws,
                    header,
                    col,
                    suffix,
                    i,
                    getattr(args, "association_context", {}),
                    getattr(args, "module_key", None),
                )
            else:
                value = minimal_field_value(header, suffix, i)
            if value not in (None, ""):
                ws.cell(row, col).value = value

        for header, col in headers.items():
            if "必填" not in header or ws.cell(row, col).value not in (None, ""):
                continue
            option = first_option_from_sheet(wb, header)
            ws.cell(row, col).value = option if option not in (None, "") else ws.cell(8, col).value

    wb.save(out)
    print(f"测试导入文件已生成: {out}")
    print(f"sheet={ws.title}, header_row={header_row}, data_start={data_start}, rows={args.rows}, fill_mode={args.fill_mode}")


def upload_to_oss(api, token, file_path):
    user_id, org_id, user = current_user(api, token)
    status, data = request_json(
        "GET",
        f"{api.rstrip('/')}{upload_to_oss.qiniu_token_path}?policy=attachment",
        token=token,
        timeout=30,
    )
    if status != 200 or "uptoken" not in data:
        raise RuntimeError(f"获取 OSS token 失败: HTTP {status} {data}")

    uptoken = data["uptoken"]
    name = Path(file_path).name
    content_type = mimetypes.guess_type(name)[0] or "application/octet-stream"
    with open(file_path, "rb") as f:
        content = f.read()

    resp = requests.post(
        uptoken["host"],
        data={
            "name": name,
            "chunk": "0",
            "chunks": "1",
            "key": str(uuid.uuid4()),
            "policy": uptoken["policy"],
            "OSSAccessKeyId": uptoken["accessid"],
            "signature": uptoken["signature"],
            "callback": uptoken.get("callback", ""),
            "success_action_status": "200",
            "Content-Disposition": "inline",
            "x:userid": user_id,
            "x:orgid": org_id,
            "x:user_token": token,
            "x:name": name,
            "x:custom_name": name,
        },
        files={"file": (name, content, content_type)},
        timeout=120,
    )
    payload = resp.json().get("payload", {})
    if not payload.get("id") or not payload.get("file_url"):
        raise RuntimeError(f"OSS 上传未返回附件信息: HTTP {resp.status_code} {resp.text[:1000]}")
    return {
        "attachment_id": payload["id"],
        "file_url": payload["file_url"],
        "user_id": user_id,
        "organization_id": org_id,
        "user_name": user.get("name"),
    }


class FayeListener:
    def __init__(self, faye_url, channel):
        self.faye_url = faye_url
        self.channel = channel
        self.session = requests.Session()
        self.messages = []
        self._stop = False
        self._thread = None

    def start(self):
        hs = self.session.post(
            self.faye_url,
            json=[{"channel": "/meta/handshake", "version": "1.0", "supportedConnectionTypes": ["long-polling"], "id": "1"}],
            timeout=15,
        ).json()[0]
        if not hs.get("successful"):
            raise RuntimeError(f"faye handshake 失败: {hs}")
        self.client_id = hs["clientId"]
        sub = self.session.post(
            self.faye_url,
            json=[{"channel": "/meta/subscribe", "clientId": self.client_id, "subscription": self.channel, "id": "2"}],
            timeout=15,
        ).json()[0]
        if not sub.get("successful"):
            raise RuntimeError(f"faye subscribe 失败: {sub}")

        def loop():
            msg_id = 3
            while not self._stop:
                try:
                    arr = self.session.post(
                        self.faye_url,
                        json=[
                            {
                                "channel": "/meta/connect",
                                "clientId": self.client_id,
                                "connectionType": "long-polling",
                                "id": str(msg_id),
                            }
                        ],
                        timeout=50,
                    ).json()
                    msg_id += 1
                    for msg in arr:
                        if msg.get("channel") == self.channel:
                            self.messages.append(msg)
                            print_json("faye_message", msg)
                except Exception:
                    pass

        self._thread = threading.Thread(target=loop, daemon=True)
        self._thread.start()

    def wait_for(self, process_type, seconds):
        deadline = time.time() + seconds
        while time.time() < deadline:
            for msg in self.messages:
                if (msg.get("data") or {}).get("process_type") == process_type:
                    return msg
            time.sleep(1)
        return None

    def stop(self):
        self._stop = True


def import_flow(args):
    resolve_business_type(args)
    file_path = Path(args.file)
    if not file_path.exists():
        raise FileNotFoundError(file_path)

    upload_to_oss.qiniu_token_path = args.qiniu_token_path
    upload = upload_to_oss(args.api, args.token, file_path)
    print_json("oss_upload", upload)

    status, new_data = request_json(
        "GET",
        entity_loader_url(args, "new"),
        token=args.token,
        params={
            "loader_name": args.loader_name,
            "custom_field_template_id": getattr(args, "custom_field_template_id", ""),
        },
        timeout=30,
    )
    new_body = new_data.get("data") or {}
    print_json(
        "entity_loaders/new",
        {
            "http_status": status,
            "code": new_data.get("code"),
            "import_client_id": new_body.get("import_client_id"),
            "import_max_count": new_body.get("import_max_count"),
            "faye_channel": new_body.get("faye_channel"),
            "duplicate_fields": new_body.get("duplicate_fields", []),
            "recognize_fields_count": len(new_body.get("recognize_fields", [])),
            "custom_field_template_id": getattr(args, "custom_field_template_id", ""),
        },
    )
    client_id = (new_data.get("data") or {}).get("import_client_id")
    channel = (new_data.get("data") or {}).get("faye_channel")
    if not client_id:
        raise RuntimeError("new 接口未返回 import_client_id")

    listener = None
    if args.listen_faye and channel:
        listener = FayeListener(args.faye_url, channel)
        listener.start()
        print(f"已监听 faye: {channel}")

    params = {
        "loader_name": args.loader_name,
        "import_client_id": client_id,
        "entity_loader_file": upload["file_url"],
        "entity_loader_file_path": "",
        "file_attachment_id": upload["attachment_id"],
        "is_overwrite": "false",
        "loggable_type": "",
        "total_row": "",
        "custom_field_template_id": getattr(args, "custom_field_template_id", ""),
    }
    status, upload_data = request_json(
        "POST", entity_loader_url(args, "upload"), token=args.token, data=params, timeout=60
    )
    print_json("entity_loaders/upload", {"http_status": status, "body": upload_data})

    validated = listener.wait_for("validated_result", args.validate_wait) if listener else None
    if listener and not validated:
        print(f"等待 {args.validate_wait}s 后未收到 validated_result")

    import_params = dict(params)
    import_params["file_attachment_id"] = upload["attachment_id"]
    import_params["entity_loader_file_path"] = upload["file_url"]
    import_params["total_row"] = str(args.rows) if hasattr(args, "rows") else ""
    status, import_data = request_json(
        "POST", entity_loader_url(args, "import"), token=args.token, data=import_params, timeout=60
    )
    print_json("entity_loaders/import", {"http_status": status, "body": import_data})

    imported = listener.wait_for("imported_result", args.import_wait) if listener else None
    import_result = {
        "loader_name": args.loader_name,
        "attachment_id": upload["attachment_id"],
        "status": "unknown",
        "total_count": None,
        "success_count": None,
        "fail_count": None,
        "error_file": None,
        "result_display": None,
    }
    if imported:
        imported_data = imported.get("data") or {}
        import_result.update(
            {
                "status": "success" if not imported_data.get("has_error_data") else "failed",
                "total_count": imported_data.get("total_count"),
                "success_count": imported_data.get("total_count") if not imported_data.get("has_error_data") else 0,
                "fail_count": (imported_data.get("error_info") or {}).get("fail_count"),
                "error_file": (imported_data.get("error_info") or {}).get("file_path"),
                "result_display": imported_data.get("result_display"),
            }
        )
    if listener and not imported:
        print(f"等待 {args.import_wait}s 后未收到 imported_result")
        import_result["status"] = "timeout"
    if listener:
        listener.stop()

    history = query_history(args)
    latest = (((history.get("body") or {}).get("data") or {}).get("list") or [{}])[0] if history else {}
    if latest.get("attachment_id") == upload["attachment_id"]:
        import_result.update(
            {
                "status": "success" if latest.get("fail_count") == 0 and latest.get("success_count", 0) > 0 else "failed",
                "total_count": latest.get("total_count"),
                "success_count": latest.get("success_count"),
                "fail_count": latest.get("fail_count"),
                "error_file": latest.get("file_path"),
                "result_display": latest.get("result_display"),
            }
        )
    query_latest_leads(args)
    print_json("import_result_summary", import_result)
    return import_result


def query_history(args):
    status, data = request_json(
        "GET",
        entity_loader_url(args, "import_histories"),
        token=args.token,
        params={"importable_type": args.loader_name, "page": 1, "per_page": args.per_page},
        timeout=30,
    )
    result = {"http_status": status, "body": data}
    print_json("import_histories", result)
    return result


def query_latest_leads(args):
    module_config = MODULE_CONFIGS.get(args.module_key, MODULE_CONFIGS["lead"])
    api_path = module_config["api_path"]
    status, data = request_json(
        "GET",
        f"{args.api.rstrip('/')}/api/v2/{api_path}",
        token=args.token,
        params={"per_page": args.per_page, "sort": "created_at", "order": "desc", "page": 1},
        timeout=30,
    )
    body = data.get("data") or {} if isinstance(data, dict) else {}
    items = body.get(api_path) or body.get("list") or []
    summary = [
        {
            "id": item.get("id"),
            "name": item.get("name") or item.get("title") or item.get("sn") or item.get("product_name"),
            "company_name": item.get("company_name") or item.get("customer_name"),
            "created_at": item.get("created_at"),
        }
        for item in items
    ]
    print_json(f"latest_{api_path}", {"http_status": status, "items": summary})


def first_item(args, module_key):
    config = MODULE_CONFIGS[module_key]
    path = config["api_path"]
    items = []
    for attempt in range(3):
        status, data = request_json(
            "GET",
            f"{args.api.rstrip('/')}/api/v2/{path}",
            token=args.token,
            params={"per_page": 1, "sort": "created_at", "order": "desc", "page": 1},
            timeout=30,
        )
        body = data.get("data") or {} if isinstance(data, dict) else {}
        items = body.get(path) or body.get("list") or []
        if status == 200 and items:
            break
        time.sleep(attempt + 1)
    if not items:
        return {}
    item = items[0]
    return {
        "id": item.get("id"),
        "name": item.get("name") or item.get("title") or item.get("contract_name") or item.get("product_name"),
        "no": item.get("product_no") or item.get("no") or item.get("sn"),
        "customer_id": item.get("customer_id"),
        "customer_name": item.get("customer_name") or item.get("customer.name"),
        "raw": item,
    }


def build_association_context(args):
    context = {
        "customer": first_item(args, "customer"),
        "contact": first_item(args, "contact"),
        "opportunity": first_item(args, "opportunity"),
        "contract": first_item(args, "contract"),
        "product": first_item(args, "product"),
    }
    if args.module_key in ("payment-plan", "received-payment", "invoiced-payment"):
        contract = context.get("contract") or {}
        if not contract.get("id"):
            raise RuntimeError("未找到可用于导入的合同，回款/开票模块需要先存在一条合同数据")
        if contract.get("customer_id"):
            context["customer"] = {
                "id": contract.get("customer_id"),
                "name": contract.get("customer_name"),
            }
    print_json(
        "association_context",
        {key: {k: v for k, v in value.items() if k != "raw"} for key, value in context.items()},
    )
    return context


def resolve_config(args):
    env_name = getattr(args, "env", None) or os.environ.get("WWJ_IMPORT_ENV") or "test"
    if env_name not in ENV_CONFIGS:
        raise SystemExit(f"未知环境: {env_name}，可选: {', '.join(ENV_CONFIGS)}")

    config = ENV_CONFIGS[env_name]
    args.env = env_name
    args.module_key = normalize_module(getattr(args, "module", None))
    args.module_config = MODULE_CONFIGS[args.module_key]
    if hasattr(args, "loader_name") and not args.loader_name:
        args.loader_name = args.module_config["loader_name"]
    args.api = getattr(args, "api", None) or os.environ.get("WWJ_IMPORT_API") or config["api"]
    args.token = getattr(args, "token", None) or os.environ.get("WWJ_USER_TOKEN")
    if args.token and args.token.startswith("user_token="):
        args.token = args.token.split("=", 1)[1]
    if hasattr(args, "entity_loader_prefix"):
        args.entity_loader_prefix = (
            args.entity_loader_prefix
            or os.environ.get("WWJ_ENTITY_LOADER_PREFIX")
            or config["entity_loader_prefix"]
        )
    if hasattr(args, "entity_loader_new_prefix"):
        args.entity_loader_new_prefix = (
            args.entity_loader_new_prefix
            or os.environ.get("WWJ_ENTITY_LOADER_NEW_PREFIX")
            or config["entity_loader_new_prefix"]
        )
    if hasattr(args, "entity_loader_history_prefix"):
        args.entity_loader_history_prefix = (
            args.entity_loader_history_prefix
            or os.environ.get("WWJ_ENTITY_LOADER_HISTORY_PREFIX")
            or config["entity_loader_history_prefix"]
        )
    if hasattr(args, "qiniu_token_path"):
        args.qiniu_token_path = (
            args.qiniu_token_path
            or os.environ.get("WWJ_QINIU_TOKEN_PATH")
            or config["qiniu_token_path"]
        )
    if hasattr(args, "faye_url"):
        args.faye_url = args.faye_url or os.environ.get("WWJ_FAYE_URL") or config["faye_url"]
    if hasattr(args, "business_type_id"):
        args.business_type_id = args.business_type_id or os.environ.get("WWJ_BUSINESS_TYPE_ID")
    if hasattr(args, "business_type_name"):
        args.business_type_name = args.business_type_name or os.environ.get("WWJ_BUSINESS_TYPE_NAME")
    if hasattr(args, "business_type_map"):
        args.business_type_map = args.business_type_map or os.environ.get("WWJ_BUSINESS_TYPE_MAP")

    if args.command != "generate-source-from-template" and not args.token:
        raise SystemExit("缺少 token：请传 --token，或设置环境变量 WWJ_USER_TOKEN")

    if args.command != "generate-source-from-template":
        print(f"使用环境: {args.env}, api={args.api}")


def full_flow(args):
    out_dir = Path(args.output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    module_name = args.module_key
    template = out_dir / f"{module_name}-template.xlsx"
    source = out_dir / f"{module_name}-source.xlsx"

    args.output = str(template)
    download_template(args)

    args.template = str(template)
    args.output = str(source)
    args.association_context = build_association_context(args) if args.fill_mode == "all" else {}
    generate_source_from_template(args)

    args.file = str(source)
    return import_flow(args)


def parse_modules(value):
    if not value or value == "all":
        return list(MODULE_CONFIGS)
    return [normalize_module(item.strip()) for item in value.split(",") if item.strip()]


def batch_full_flow(args):
    results = []
    for module_key in parse_modules(args.modules):
        child = argparse.Namespace(**vars(args))
        child.module = module_key
        child.module_key = module_key
        child.module_config = MODULE_CONFIGS[module_key]
        child.loader_name = child.module_config["loader_name"]
        child.output_dir = str(Path(args.output_dir) / module_key)
        print(f"\n===== 开始处理模块: {module_key} / {child.module_config['label']} =====")
        try:
            result = full_flow(child)
            results.append({"module": module_key, "label": child.module_config["label"], **(result or {})})
        except Exception as exc:
            result = {"module": module_key, "label": child.module_config["label"], "status": "failed", "error": str(exc)}
            results.append(result)
            print_json("module_failed", result)
            if not args.continue_on_error:
                raise
    print_json("batch_summary", results)


def build_parser():
    parser = argparse.ArgumentParser(description="CRM 导入完整流程测试工具")
    common = argparse.ArgumentParser(add_help=False)
    common.add_argument("--env", choices=sorted(ENV_CONFIGS), default=None, help="预置环境，默认 test")
    common.add_argument("--api", default=None, help="覆盖环境域名")
    common.add_argument("--token", default=None, help="user_token；也可用 WWJ_USER_TOKEN")
    common.add_argument("--module", default="lead", help="业务模块，例如 lead/customer/contact/opportunity")
    common.add_argument("--loader-name", default=None, help="覆盖模块对应的 loader_name")
    common.add_argument("--entity-loader-prefix", default=None, help="覆盖导入接口前缀")
    common.add_argument("--entity-loader-new-prefix", default=None, help="覆盖 new 接口前缀")
    common.add_argument("--entity-loader-history-prefix", default=None, help="覆盖导入历史接口前缀")
    common.add_argument("--qiniu-token-path", default=None, help="覆盖 OSS token 接口路径")
    common.add_argument("--business-type-id", default=None, help="客户/商机/报价单/合同导入使用的业务类型 ID")
    common.add_argument("--business-type-name", default=None, help="客户/商机/报价单/合同导入使用的业务类型名称")
    common.add_argument(
        "--business-type-map",
        default=None,
        help="批量时按模块指定业务类型，例如 customer=793,opportunity=商机业务类型2",
    )

    sub = parser.add_subparsers(dest="command", required=True)

    p = sub.add_parser("download-template", parents=[common])
    p.add_argument("--output", default=str(DEFAULT_OUT_DIR / "lead-template.xlsx"))
    p.set_defaults(func=download_template)

    p = sub.add_parser("generate-source-from-template")
    p.add_argument("--template", required=True)
    p.add_argument("--output", required=True)
    p.add_argument("--rows", type=int, default=2)
    p.add_argument("--fill-mode", choices=["all", "minimal"], default="all")
    p.set_defaults(func=generate_source_from_template)

    p = sub.add_parser("import-flow", parents=[common])
    p.add_argument("--file", required=True)
    p.add_argument("--faye-url", default=None)
    p.add_argument("--listen-faye", action=argparse.BooleanOptionalAction, default=True)
    p.add_argument("--validate-wait", type=int, default=45)
    p.add_argument("--import-wait", type=int, default=90)
    p.add_argument("--per-page", type=int, default=5)
    p.set_defaults(func=import_flow)

    p = sub.add_parser("full-flow", parents=[common])
    p.add_argument("--output-dir", default=str(DEFAULT_OUT_DIR))
    p.add_argument("--rows", type=int, default=2)
    p.add_argument("--fill-mode", choices=["all", "minimal"], default="all")
    p.add_argument("--faye-url", default=None)
    p.add_argument("--listen-faye", action=argparse.BooleanOptionalAction, default=True)
    p.add_argument("--validate-wait", type=int, default=45)
    p.add_argument("--import-wait", type=int, default=90)
    p.add_argument("--per-page", type=int, default=5)
    p.set_defaults(func=full_flow)

    p = sub.add_parser("batch-full-flow", parents=[common])
    p.add_argument(
        "--modules",
        default="all",
        help="逗号分隔模块，默认 all；例如 customer,contact,opportunity",
    )
    p.add_argument("--output-dir", default=str(DEFAULT_OUT_DIR / "batch"))
    p.add_argument("--rows", type=int, default=1)
    p.add_argument("--fill-mode", choices=["all", "minimal"], default="all")
    p.add_argument("--faye-url", default=None)
    p.add_argument("--listen-faye", action=argparse.BooleanOptionalAction, default=True)
    p.add_argument("--validate-wait", type=int, default=45)
    p.add_argument("--import-wait", type=int, default=90)
    p.add_argument("--per-page", type=int, default=3)
    p.add_argument("--continue-on-error", action=argparse.BooleanOptionalAction, default=True)
    p.set_defaults(func=batch_full_flow)

    p = sub.add_parser("business-types", parents=[common])
    p.set_defaults(func=print_business_types)

    p = sub.add_parser("history", parents=[common])
    p.add_argument("--per-page", type=int, default=5)
    p.set_defaults(func=query_history)
    return parser


def main():
    args = build_parser().parse_args()
    resolve_config(args)
    args.func(args)


if __name__ == "__main__":
    main()
