#!/usr/bin/env python3
"""
CRM Excel import helper for Lead/error-report import testing.

常用用法:
  环境配置:
     --env test 读取 config.test.json
     --env staging 读取 config.staging.json
     --env staging --profile gray 读取 config.staging.gray.json 和 import/staging/gray/
     token 失效时可追加 --token '<新token>' 临时覆盖配置

  1. 下载模板
     python3 error_report_import_tester.py download-template --module lead --output lead-template.xlsx

  2. 按固定条数生成导入源文件
     python3 error_report_import_tester.py generate-source-from-template \
       --template lead-template.xlsx --output lead-source.xlsx --rows 3

  3. 按随机条数生成导入源文件
     python3 error_report_import_tester.py generate-source-from-template \
       --template lead-template.xlsx --output lead-source.xlsx --rows-random 1 10

  4. 全流程执行
     python3 /Users/mured/wwj/wwjapi/import/error-report-import/error_report_import_tester.py full-flow \
       --env test \
       --module contract \
       --rows 10

     测试环境页面下载的系统模板可放在以下目录，脚本会优先按表头自动匹配:
       /Users/mured/wwj/wwjapi/import/test/
       /Users/mured/wwj/wwjapi/templates/import/test/

  5. 随机条数全流程执行
     python3 error_report_import_tester.py full-flow --module lead --rows-random 1 10

  6. 线索池导入
     python3 error_report_import_tester.py full-flow --module lead-pool --common-id 5980 --rows 5

  7. 错误模板二次导入流程
     python3 /Users/mured/wwj/wwjapi/import/error-report-import/error_report_import_tester.py error-template-flow \
       --env test \
       --module lead \
       --rows 1

     python3 /Users/mured/wwj/wwjapi/import/error-report-import/error_report_import_tester.py error-template-flow \
       --env test \
       --module lead-pool \
       --common-id 5980 \
       --rows 5

     python3 /Users/mured/wwj/wwjapi/import/error-report-import/error_report_import_tester.py error-template-flow \
       --env test \
       --module customer-pool \
       --common-id 11033 \
       --rows 5
"""

import argparse
import hashlib
import json
import mimetypes
import os
import random
import re
import shutil
import subprocess
import sys
import tempfile
import threading
import time
import uuid
from copy import deepcopy
from datetime import datetime, timedelta
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile
from xml.etree import ElementTree as ET

import requests
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation, DataValidationList

PROJECT_ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(PROJECT_ROOT))
from lib.config import load_config


DEFAULT_OUT_DIR = Path("outputs/error-report-import")
DEFAULT_FAYE_URL = "https://faye-dev.ikcrm.com/faye"
DEFAULT_DATA_RELATION_CUSTOM_FORM_ID = "1125"
SHEET_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
PKG_CT_NS = "http://schemas.openxmlformats.org/package/2006/content-types"
PKG_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
OFFICE_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
ET.register_namespace("", SHEET_NS)
ET.register_namespace("rel", PKG_REL_NS)
ET.register_namespace("ct", PKG_CT_NS)
DEFAULT_CUSTOM_PROPS_XML = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Properties xmlns="http://schemas.openxmlformats.org/officeDocument/2006/custom-properties" xmlns:vt="http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes">
  <property fmtid="{D5CDD505-2E9C-101B-9397-08002B2CF9AE}" pid="2" name="ICV"><vt:lpwstr>5E17FFB40147427028104E6AC30EC769_42</vt:lpwstr></property>
  <property fmtid="{D5CDD505-2E9C-101B-9397-08002B2CF9AE}" pid="3" name="KSOProductBuildVer"><vt:lpwstr>2052-12.1.26026.26026</vt:lpwstr></property>
  <property fmtid="{D5CDD505-2E9C-101B-9397-08002B2CF9AE}" pid="4" name="CalculationRule"><vt:i4>0</vt:i4></property>
</Properties>
""".encode("utf-8")
MODULE_CONFIGS = {
    "lead": {"loader_name": "Lead", "api_path": "leads", "label": "线索"},
    "lead-pool": {"loader_name": "LeadCommon", "api_path": "leads", "label": "线索池"},
    "customer": {"loader_name": "Customer", "api_path": "customers", "label": "客户"},
    "customer-pool": {"loader_name": "CustomerCommon", "api_path": "customers", "label": "客户公海"},
    "contact": {"loader_name": "Contact", "api_path": "contacts", "label": "联系人"},
    "opportunity": {"loader_name": "Opportunity", "api_path": "opportunities", "label": "商机"},
    "quotation": {"loader_name": "Quotation", "api_path": "quotations", "label": "报价单"},
    "contract": {"loader_name": "Contract", "api_path": "contracts", "label": "合同"},
    "payment-plan": {"loader_name": "ReceivedPaymentPlan", "api_path": "received_payment_plans", "label": "回款计划"},
    "received-payment": {"loader_name": "ReceivedPayment", "api_path": "received_payments", "label": "回款记录"},
    "invoiced-payment": {"loader_name": "InvoicedPayment", "api_path": "invoiced_payments", "label": "开票记录"},
    "product": {"loader_name": "Product", "api_path": "products", "label": "产品"},
}
MODULE_ALIASES = {
    "线索": "lead",
    "线索池": "lead-pool",
    "客户": "customer",
    "客户公海": "customer-pool",
    "联系人": "contact",
    "商机": "opportunity",
    "报价单": "quotation",
    "合同": "contract",
    "回款计划": "payment-plan",
    "回款记录": "received-payment",
    "开票记录": "invoiced-payment",
    "产品": "product",
}
BUSINESS_TYPE_MODULES = {"customer", "opportunity", "quotation", "contract"}
ASSOCIATION_CONTEXT_MODULES = {
    "lead",
    "lead-pool",
    "customer",
    "contact",
    "opportunity",
    "quotation",
    "contract",
    "payment-plan",
    "received-payment",
    "invoiced-payment",
}
ENV_CONFIGS = {
    "test": {
        "api": "https://lxcrm-test.weiwenjia.com",
        "faye_url": DEFAULT_FAYE_URL,
        "entity_loader_prefix": "/api/pc/entity_loaders",
        "entity_loader_new_prefix": "/api/pc/entity_loaders",
        "entity_loader_history_prefix": "/api/pc/entity_loaders",
        "qiniu_token_path": "/api/pc/qiniu/auth/oss_upload_token.json",
    },
    "staging": {
        "api": "https://lxcrm-staging.weiwenjia.com",
        "faye_url": DEFAULT_FAYE_URL,
        "entity_loader_prefix": "/api/entity_loaders",
        "entity_loader_new_prefix": "/api/pc/entity_loaders",
        "entity_loader_history_prefix": "/api/pc/entity_loaders",
        "qiniu_token_path": "/api/qiniu/auth/oss_upload_token.json",
    },
    "production": {
        "api": "https://lxcrm.weiwenjia.com",
        "faye_url": DEFAULT_FAYE_URL,
        "entity_loader_prefix": "/api/entity_loaders",
        "entity_loader_new_prefix": "/api/pc/entity_loaders",
        "entity_loader_history_prefix": "/api/pc/entity_loaders",
        "qiniu_token_path": "/api/qiniu/auth/oss_upload_token.json",
    },
}


def pc_url(api):
    return api.rstrip("/").replace("//lxcrm-test.", "//lxcrm-api-test.").replace(
        "//lxcrm-staging.", "//lxcrm-api-staging."
    ).replace("//lxcrm.", "//lxcrm-api.")


def auth_headers(token):
    return {
        "Accept": "application/json, text/plain, */*",
        "Authorization": f'Token token="{token}",device="web"',
        "ACCESS-TOKEN": token,
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/149.0.0.0 Safari/537.36",
    }


def print_json(title, data):
    print(f"\n## {title}")
    print(json.dumps(data, ensure_ascii=False, indent=2))


def file_sha1(path):
    h = hashlib.sha1()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def remote_file_digest(url):
    resp = requests.get(url, timeout=60)
    resp.raise_for_status()
    return {"size": len(resp.content), "sha1": hashlib.sha1(resp.content).hexdigest()}


def sheet_q(tag):
    return f"{{{SHEET_NS}}}{tag}"


def has_preservable_template_package(path):
    try:
        with ZipFile(path) as zf:
            names = set(zf.namelist())
            return "xl/workbook.xml" in names and "xl/worksheets/sheet1.xml" in names
    except Exception:
        return False


def template_signature(path):
    try:
        wb = load_workbook(path, data_only=False, read_only=True)
        ws = wb.active
        header_row, _ = find_template_rows(ws)
        headers = tuple(str(ws.cell(header_row, col).value or "") for col in range(1, ws.max_column + 1))
        sheets = tuple(wb.sheetnames)
        return headers, sheets
    except Exception:
        return None


def template_headers(path):
    signature = template_signature(path)
    return signature[0] if signature else None


def template_search_dirs(args):
    dirs = []
    env = str(getattr(args, "env", "test") or "test")
    profile = getattr(args, "profile", None)
    for raw in (
        getattr(args, "template_dir", None),
        os.environ.get("WWJ_IMPORT_TEMPLATE_DIR"),
        Path.cwd() / "import" / env / profile if profile else None,
        Path.cwd() / "templates" / "import" / env / profile if profile else None,
        Path.cwd() / "import" / env,
        Path.cwd() / "templates" / "import" / env,
        Path.cwd() / "templates" / "import",
        Path.home() / "Downloads",
        Path("/Users/mured/Downloads"),
    ):
        if not raw:
            continue
        path = Path(raw)
        if path.exists() and path.is_dir() and path not in dirs:
            dirs.append(path)
    return dirs


def local_browser_template(args, api_template=None):
    explicit = getattr(args, "template_file", None)
    if explicit:
        return Path(explicit)

    name = f"CRM_{args.module_config['label']}_导入模板.xlsx"
    dirs = template_search_dirs(args)
    exact_candidates = [directory / name for directory in dirs]
    api_headers = template_headers(api_template) if api_template else None
    if api_headers:
        existing_exact_candidates = [item for item in exact_candidates if item.exists()]
        for item in existing_exact_candidates:
            if template_headers(item) == api_headers:
                return item

        # A same-name template belongs to this module. If its headers differ
        # from the current API template, it is stale and must not be reused.
        if existing_exact_candidates:
            return None

        candidates = []
        for directory in dirs:
            candidates.extend(directory.glob("CRM_*_导入模板.xlsx"))
        matched = [item for item in candidates if template_headers(item) == api_headers]
        if matched:
            return max(matched, key=lambda item: item.stat().st_mtime)
        return None

    for item in exact_candidates:
        if item.exists():
            return item
    return None


def matching_local_business_template(args):
    if args.module_key not in BUSINESS_TYPE_MODULES or business_type_value(args):
        return None

    local_templates = []
    for directory in template_search_dirs(args):
        local_templates.extend(directory.glob("CRM_*_导入模板.xlsx"))
    if not local_templates:
        return None

    for item in list_business_types(args):
        if item.get("status") != "enable":
            continue
        response = requests.get(
            entity_loader_url(args, f"template/{args.loader_name}.xlsx"),
            headers=auth_headers(args.token),
            params={"custom_field_template_id": item["id"]},
            timeout=60,
        )
        if response.status_code != 200:
            continue

        fd, remote_name = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        remote_path = Path(remote_name)
        try:
            remote_path.write_bytes(response.content)
            remote_headers = template_headers(remote_path)
        finally:
            remote_path.unlink(missing_ok=True)
        if not remote_headers:
            continue

        for local_template in local_templates:
            if template_headers(local_template) != remote_headers:
                continue
            args.custom_field_template_id = str(item["id"])
            args.business_type_resolved = True
            print(f"业务类型: {args.module_config['label']} -> {item.get('name')}({item.get('id')})")
            return local_template
    return None


def compact_data_validations(ws):
    grouped = {}
    for dv in list(ws.data_validations.dataValidation):
        key = (
            dv.type,
            dv.formula1,
            dv.formula2,
            dv.allow_blank,
            dv.operator,
            dv.showDropDown,
            dv.showErrorMessage,
            dv.showInputMessage,
            dv.errorStyle,
            dv.errorTitle,
            dv.error,
            dv.promptTitle,
            dv.prompt,
        )
        for cell_range in dv.sqref.ranges:
            if cell_range.min_col != cell_range.max_col:
                grouped.setdefault((key, str(cell_range)), []).append(cell_range)
                continue
            col = cell_range.min_col
            grouped.setdefault((key, col), []).append(cell_range)

    if not grouped:
        return

    ws.data_validations = DataValidationList()
    for (key, col_or_range), ranges in grouped.items():
        (
            type_,
            formula1,
            formula2,
            allow_blank,
            operator,
            show_drop_down,
            show_error_message,
            show_input_message,
            error_style,
            error_title,
            error,
            prompt_title,
            prompt,
        ) = key
        new_dv = DataValidation(
            type=type_,
            formula1=formula1,
            formula2=formula2,
            allow_blank=allow_blank,
            operator=operator,
            showDropDown=show_drop_down,
            showErrorMessage=show_error_message,
            showInputMessage=show_input_message,
            errorStyle=error_style,
            errorTitle=error_title,
            error=error,
            promptTitle=prompt_title,
            prompt=prompt,
        )
        if isinstance(col_or_range, str):
            new_dv.add(col_or_range)
        else:
            min_row = min(item.min_row for item in ranges)
            max_row = max(item.max_row for item in ranges)
            col_letter = get_column_letter(col_or_range)
            new_dv.add(f"{col_letter}{min_row}:{col_letter}{max_row}")
        ws.add_data_validation(new_dv)


def read_shared_strings(data):
    if "xl/sharedStrings.xml" not in data:
        return None, [], []
    root = ET.fromstring(data["xl/sharedStrings.xml"])
    strings = []
    items = root.findall(sheet_q("si"))
    for si in items:
        strings.append("".join(t.text or "" for t in si.iter(sheet_q("t"))))
    return root, items, strings


def ensure_content_type_override(types_root, part_name, content_type):
    for child in types_root.findall(f"{{{PKG_CT_NS}}}Override"):
        if child.get("PartName") == part_name:
            return
    ET.SubElement(types_root, f"{{{PKG_CT_NS}}}Override", {"PartName": part_name, "ContentType": content_type})


def ensure_workbook_relationship(rels_root, rel_type, target):
    for child in rels_root.findall(f"{{{PKG_REL_NS}}}Relationship"):
        if child.get("Type") == rel_type or child.get("Target") == target:
            return
    ids = []
    for child in rels_root.findall(f"{{{PKG_REL_NS}}}Relationship"):
        rid = child.get("Id", "")
        match = re.match(r"rId(\\d+)", rid)
        if match:
            ids.append(int(match.group(1)))
    next_id = f"rId{max(ids) + 1 if ids else 1}"
    ET.SubElement(rels_root, f"{{{PKG_REL_NS}}}Relationship", {"Id": next_id, "Type": rel_type, "Target": target})


def ensure_package_metadata(data, shared_strings_created=False):
    content_types = ET.fromstring(data["[Content_Types].xml"])
    ensure_content_type_override(
        content_types,
        "/docProps/custom.xml",
        "application/vnd.openxmlformats-officedocument.custom-properties+xml",
    )
    if shared_strings_created:
        ensure_content_type_override(
            content_types,
            "/xl/sharedStrings.xml",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml",
        )
    data["[Content_Types].xml"] = ET.tostring(content_types, encoding="utf-8", xml_declaration=True)

    if shared_strings_created:
        rels = ET.fromstring(data["xl/_rels/workbook.xml.rels"])
        ensure_workbook_relationship(
            rels,
            "http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings",
            "sharedStrings.xml",
        )
        data["xl/_rels/workbook.xml.rels"] = ET.tostring(rels, encoding="utf-8", xml_declaration=True)
    if "docProps/custom.xml" not in data:
        data["docProps/custom.xml"] = DEFAULT_CUSTOM_PROPS_XML


def write_shared_string(shared_root, shared_items, strings, text):
    if text in strings:
        return strings.index(text)
    si = ET.SubElement(shared_root, sheet_q("si"))
    t = ET.SubElement(si, sheet_q("t"))
    if str(text).strip() != str(text):
        t.set("{http://www.w3.org/XML/1998/namespace}space", "preserve")
    t.text = str(text)
    shared_items.append(si)
    strings.append(str(text))
    return len(strings) - 1


def convert_inline_strings_to_shared(sheet, shared_root, shared_items, strings):
    sheet_data = sheet.find(sheet_q("sheetData"))
    if sheet_data is None:
        return
    for row_el in sheet_data.findall(sheet_q("row")):
        for cell in row_el.findall(sheet_q("c")):
            if cell.get("t") != "inlineStr" and cell.find(sheet_q("is")) is None:
                continue
            text = "".join(t.text or "" for t in cell.iter(sheet_q("t")))
            if text == "":
                continue
            for child in list(cell):
                cell.remove(child)
            cell.set("t", "s")
            v = ET.SubElement(cell, sheet_q("v"))
            v.text = str(write_shared_string(shared_root, shared_items, strings, text))


def set_sheet_cell(row_el, cell_ref, value, shared_root, shared_items, strings):
    cell = None
    for item in row_el.findall(sheet_q("c")):
        if item.get("r") == cell_ref:
            cell = item
            break
    if cell is None:
        cell = ET.SubElement(row_el, sheet_q("c"), {"r": cell_ref})
    for child in list(cell):
        cell.remove(child)
    if value in (None, ""):
        cell.attrib.pop("t", None)
        return
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        cell.attrib.pop("t", None)
        v = ET.SubElement(cell, sheet_q("v"))
        v.text = str(value)
        return
    if shared_root is None:
        cell.set("t", "inlineStr")
        is_el = ET.SubElement(cell, sheet_q("is"))
        t = ET.SubElement(is_el, sheet_q("t"))
        if str(value).strip() != str(value):
            t.set("{http://www.w3.org/XML/1998/namespace}space", "preserve")
        t.text = str(value)
        return
    cell.set("t", "s")
    v = ET.SubElement(cell, sheet_q("v"))
    v.text = str(write_shared_string(shared_root, shared_items, strings, str(value)))


def sort_sheet_row_cells(row_el):
    cells = row_el.findall(sheet_q("c"))
    if len(cells) < 2:
        return
    for cell in cells:
        row_el.remove(cell)

    def cell_col_index(cell):
        ref = cell.get("r", "")
        col = re.sub(r"\d+", "", ref)
        return column_index_from_string(col) if col else 0

    for cell in sorted(cells, key=cell_col_index):
        row_el.append(cell)


def generate_source_preserving_package(args):
    template = Path(args.template)
    out = Path(args.output)
    wb = load_workbook(template, data_only=False)
    ws = wb.active
    header_row, data_start = find_template_rows(ws)
    headers = {str(ws.cell(header_row, col).value or ""): col for col in range(1, ws.max_column + 1)}
    suffix = datetime.now().strftime("%H%M%S") + uuid.uuid4().hex[:6]
    row_values = []
    for i in range(1, args.rows + 1):
        values = {}
        for header, col in headers.items():
            if args.fill_mode == "all":
                value = full_field_value(
                    wb,
                    ws,
                    header,
                    col,
                    suffix,
                    i,
                    getattr(args, "association_context", {}),
                    getattr(args, "module_key", None),
                    args,
                )
            else:
                value = minimal_field_value(header, suffix, i)
            if args.module_key == "customer" and col == 6:
                value = None
            values[col] = value
        row_values.append(values)

    with ZipFile(template) as zin:
        data = {name: zin.read(name) for name in zin.namelist()}

    shared_root, shared_items, strings = read_shared_strings(data)
    shared_strings_created = False
    if shared_root is None:
        shared_root = ET.Element(sheet_q("sst"), {"count": "0", "uniqueCount": "0"})
        shared_items = []
        strings = []
        shared_strings_created = True
    sheet = ET.fromstring(data["xl/worksheets/sheet1.xml"])
    convert_inline_strings_to_shared(sheet, shared_root, shared_items, strings)
    sheet_data = sheet.find(sheet_q("sheetData"))
    rows_by_num = {int(row.get("r")): row for row in sheet_data.findall(sheet_q("row")) if row.get("r")}
    header_row_el = rows_by_num.get(data_start - 1)
    if header_row_el is not None and header_row_el.get("spans") is None:
        header_row_el.set("spans", f"1:{ws.max_column}")
    max_target_row = data_start + args.rows - 1

    for row_num, row_el in list(rows_by_num.items()):
        if row_num > max_target_row and row_num >= data_start:
            sheet_data.remove(row_el)
            rows_by_num.pop(row_num, None)

    template_row = rows_by_num.get(data_start)
    if template_row is None and (data_start - 1) in rows_by_num:
        template_row = rows_by_num.get(data_start - 1)
    for row_num in range(data_start, max_target_row + 1):
        row_el = rows_by_num.get(row_num)
        if row_el is None:
            row_el = deepcopy(template_row) if template_row is not None else ET.Element(sheet_q("row"))
            row_el.set("r", str(row_num))
            for cell in row_el.findall(sheet_q("c")):
                col = re.sub(r"\d+", "", cell.get("r", ""))
                cell.set("r", f"{col}{row_num}")
            sheet_data.append(row_el)
            rows_by_num[row_num] = row_el
        if row_el.get("spans") is None:
            row_el.set("spans", f"1:{ws.max_column}")
        values = row_values[row_num - data_start] if row_num - data_start < len(row_values) else {}
        for col in range(1, ws.max_column + 1):
            cell_ref = f"{get_column_letter(col)}{row_num}"
            value = values.get(col)
            if value in (None, ""):
                for cell in row_el.findall(sheet_q("c")):
                    if cell.get("r") == cell_ref:
                        row_el.remove(cell)
                        break
                continue
            set_sheet_cell(row_el, cell_ref, value, shared_root, shared_items, strings)
        sort_sheet_row_cells(row_el)

    dimension = sheet.find(sheet_q("dimension"))
    if dimension is not None:
        dimension.set("ref", f"A1:{get_column_letter(ws.max_column)}{max_target_row}")
    if shared_root is not None:
        shared_root.set("count", str(len(strings)))
        shared_root.set("uniqueCount", str(len(strings)))
        data["xl/sharedStrings.xml"] = ET.tostring(shared_root, encoding="utf-8", xml_declaration=True)
    ensure_package_metadata(data, shared_strings_created=shared_strings_created)
    data["xl/worksheets/sheet1.xml"] = ET.tostring(sheet, encoding="utf-8", xml_declaration=True)

    out.parent.mkdir(parents=True, exist_ok=True)
    with ZipFile(out, "w", ZIP_DEFLATED) as zout:
        for name, content in data.items():
            zout.writestr(name, content)
    print(f"测试导入文件已生成: {out}")
    print(
        f"sheet={ws.title}, header_row={header_row}, data_start={data_start}, rows={args.rows}, "
        f"fill_mode={args.fill_mode}, package=preserved"
    )


def normalize_module(module):
    if not module:
        return "lead"
    key = MODULE_ALIASES.get(module, module).strip().lower().replace("_", "-")
    if key not in MODULE_CONFIGS:
        choices = ", ".join(MODULE_CONFIGS)
        raise SystemExit(f"未知模块: {module}，可选: {choices}")
    return key


def request_json(method, url, *, token, **kwargs):
    headers = kwargs.pop("headers", {})
    merged = auth_headers(token)
    if "lxcrm-api-test.weiwenjia.com" in url or "lxcrm-test.weiwenjia.com" in url:
        merged.setdefault("Origin", "https://lxcrm-test.weiwenjia.com")
        merged.setdefault("Referer", "https://lxcrm-test.weiwenjia.com/")
    elif "lxcrm-api-staging.weiwenjia.com" in url or "lxcrm-staging.weiwenjia.com" in url:
        merged.setdefault("Origin", "https://lxcrm-staging.weiwenjia.com")
        merged.setdefault("Referer", "https://lxcrm-staging.weiwenjia.com/")
    elif "lxcrm-api.weiwenjia.com" in url or "lxcrm.weiwenjia.com" in url:
        merged.setdefault("Origin", "https://lxcrm.weiwenjia.com")
        merged.setdefault("Referer", "https://lxcrm.weiwenjia.com/")
    merged.setdefault("Accept-Language", "zh-CN,zh;q=0.9")
    merged.setdefault("x-lx-gid", "")
    merged.update(headers)
    resp = requests.request(method, url, headers=merged, timeout=kwargs.pop("timeout", 60), **kwargs)
    try:
        data = resp.json()
    except ValueError:
        data = {"raw_text": resp.text[:2000]}
    return resp.status_code, data


def parse_business_type_map(value):
    result = {}
    if not value:
        return result
    for item in value.split(","):
        if not item.strip():
            continue
        if "=" not in item:
            raise SystemExit(f"业务类型映射格式错误: {item}，应为 module=id 或 module=name")
        module, template = item.split("=", 1)
        result[normalize_module(module.strip())] = template.strip()
    return result


def business_type_value(args):
    mapped = parse_business_type_map(getattr(args, "business_type_map", None)).get(args.module_key)
    return mapped or getattr(args, "business_type_id", None) or getattr(args, "business_type_name", None)


def list_business_types(args):
    status, data = request_json(
        "GET",
        f"{pc_url(args.api)}/api/pc/custom_field_templates",
        token=args.token,
        params={"model_klass": args.module_key.replace("-", "_")},
        timeout=30,
    )
    items = (data.get("data") or {}).get("list") or []
    if status == 200 and items:
        return items

    model_klass = args.module_config["loader_name"]
    status, data = request_json(
        "GET",
        f"{pc_url(args.api)}/api/pc/custom_fields",
        token=args.token,
        params={"model_klass": model_klass},
        timeout=30,
    )
    groups = (data.get("data") or {}).get("custom_field_groups") or []
    field_id = None
    for group in groups:
        for field in group.get("custom_fields") or []:
            if field.get("field_id"):
                field_id = field["field_id"]
                break
        if field_id:
            break
    if status != 200 or not field_id:
        return []

    status, detail = request_json(
        "GET",
        f"{pc_url(args.api)}/api/pc/custom_fields/{field_id}",
        token=args.token,
        params={"custom_field_template_id": 1331},
        timeout=30,
    )
    if status != 200:
        return []
    return (detail.get("data") or {}).get("custom_field_templates") or []


def resolve_business_type(args):
    if getattr(args, "business_type_resolved", False):
        return
    if args.module_key not in BUSINESS_TYPE_MODULES:
        args.custom_field_template_id = ""
        args.business_type_resolved = True
        return

    requested = business_type_value(args)
    templates = list_business_types(args)
    enabled = [item for item in templates if item.get("status") == "enable"]
    selected = None
    if requested:
        for item in enabled:
            if str(item.get("id")) == str(requested) or item.get("name") == requested:
                selected = item
                break
        if not selected:
            names = ", ".join(f"{item.get('id')}:{item.get('name')}" for item in enabled)
            raise RuntimeError(f"{args.module_config['label']} 未找到业务类型 {requested}，可选: {names}")
    elif enabled:
        selected = max(enabled, key=lambda item: int(item.get("id") or 0))

    if not selected:
        args.custom_field_template_id = ""
        args.business_type_resolved = True
        print(f"{args.module_config['label']} 未发现启用业务类型，继续使用默认模板")
        return

    args.custom_field_template_id = str(selected["id"])
    args.business_type_resolved = True
    print(f"业务类型: {args.module_config['label']} -> {selected.get('name')}({selected.get('id')})")


def print_business_types(args):
    if args.module_key not in BUSINESS_TYPE_MODULES:
        print_json("business_types", {"module": args.module_key, "items": []})
        return []
    items = [
        {"id": item.get("id"), "name": item.get("name"), "status": item.get("status")}
        for item in list_business_types(args)
    ]
    print_json("business_types", {"module": args.module_key, "items": items})
    return items


def entity_loader_url(args, path):
    clean_path = path.lstrip("/")
    if clean_path == "new":
        prefix = args.entity_loader_new_prefix
    elif clean_path == "import_histories":
        prefix = args.entity_loader_history_prefix
    else:
        prefix = args.entity_loader_prefix
    return f"{pc_url(args.api)}{prefix}/{path.lstrip('/')}"


def current_user(api, token):
    status, data = request_json("GET", f"{pc_url(api)}/api/v2/user/info", token=token, timeout=30)
    if status != 200 or data.get("code") != 0:
        raise RuntimeError(f"获取当前用户失败: HTTP {status} {data}")
    info = data["data"]
    return str(info["id"]), str(info["organization_id"]), info


def download_template(args):
    resolve_business_type(args)
    out = Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)
    url = entity_loader_url(args, f"template/{args.loader_name}.xlsx")
    params = {}
    if getattr(args, "custom_field_template_id", ""):
        params["custom_field_template_id"] = args.custom_field_template_id
    if getattr(args, "common_id", "") and args.module_key in ("lead-pool", "customer-pool"):
        params["common_id"] = args.common_id
    resp = requests.get(url, headers=auth_headers(args.token), params=params, timeout=60)
    resp.raise_for_status()
    out.write_bytes(resp.content)
    print(f"模板已下载: {out} ({len(resp.content)} bytes)")


def find_template_rows(ws):
    marker_row = None
    for row in range(1, min(ws.max_row, 80) + 1):
        first = ws.cell(row, 1).value
        if isinstance(first, str) and "请从下面表格开始填写" in first:
            marker_row = row
            break
    if marker_row:
        return marker_row + 1, marker_row + 2

    # Fallback for older templates.
    for row in range(1, min(ws.max_row, 80) + 1):
        values = [ws.cell(row, col).value for col in range(1, min(ws.max_column, 30) + 1)]
        if any(isinstance(v, str) and "必填" in v for v in values):
            return row, row + 1
    raise RuntimeError("无法识别模板表头行")


def clean_header(header):
    return str(header or "").replace("(必填)", "").replace("（必填）", "").strip()


def first_option_from_sheet(wb, header):
    values = option_values_from_sheet(wb, header, limit=1)
    return values[0] if values else None


def option_values_from_sheet(wb, header, limit=2, example=None):
    name = clean_header(header)
    stripped_name = re.sub(r"[\[\]（）()]", "", name).strip()
    candidates = [f"Sheet{name}", f"Sheet_{name}"]
    if stripped_name and stripped_name != name:
        candidates.extend([f"Sheet{stripped_name}", f"Sheet_{stripped_name}"])
    short_name = None
    if "]" in name:
        short_name = name.rsplit("]", 1)[-1].strip()
        if short_name:
            candidates.extend([f"Sheet{short_name}", f"Sheet_{short_name}"])
    exact_candidates = list(dict.fromkeys(candidates))
    if isinstance(example, str):
        candidates.extend(re.findall(r"Sheet[-_\w\u4e00-\u9fff]+", example))
    for ws in wb.worksheets:
        if ws.title in exact_candidates:
            values = []
            for row in range(2, ws.max_row + 1):
                value = ws.cell(row, 1).value
                if value not in (None, ""):
                    values.append(value)
                if len(values) >= limit:
                    break
            return values
    normalized = [c.replace("_", "").lower() for c in candidates]
    for ws in wb.worksheets:
        title = ws.title.replace("_", "").lower()
        if ws.title in candidates or title in normalized:
            values = []
            for row in range(2, ws.max_row + 1):
                value = ws.cell(row, 1).value
                if value not in (None, ""):
                    values.append(value)
                if len(values) >= limit:
                    break
            return values
    if stripped_name:
        for ws in wb.worksheets:
            title = ws.title.replace("_", "").lower()
            if stripped_name.replace("_", "").lower() in title:
                values = []
                for row in range(2, ws.max_row + 1):
                    value = ws.cell(row, 1).value
                    if value not in (None, ""):
                        values.append(value)
                    if len(values) >= limit:
                        break
                return values
    if short_name:
        for ws in wb.worksheets:
            if short_name.replace("_", "").lower() in ws.title.replace("_", "").lower():
                values = []
                for row in range(2, ws.max_row + 1):
                    value = ws.cell(row, 1).value
                    if value not in (None, ""):
                        values.append(value)
                    if len(values) >= limit:
                        break
                return values
    return []


def user_option_values(wb, limit=2):
    sheet_name = "Sheet自定义用户"
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    values = []
    for row in range(2, ws.max_row + 1):
        value = ws.cell(row, 1).value
        if value not in (None, ""):
            values.append(value)
        if len(values) >= limit:
            return values
    return values


def pc_user_name_values(args, limit=50):
    cache = getattr(args, "_pc_user_cache", None)
    if cache is None:
        try:
            status, data = request_json(
                "GET",
                f"{pc_url(args.api)}/api/pc/users",
                token=args.token,
                params={"page": 1, "per_page": limit},
                timeout=30,
            )
            if status == 200 and data.get("code") == 0:
                body = data.get("data") or {}
                users = body.get("users") or body.get("list") or []
                cache = [u.get("name") for u in users if u.get("name")]
            else:
                cache = []
        except requests.RequestException:
            cache = []
        args._pc_user_cache = cache
    return cache or []


def selected_user_values(wb, args=None, limit=2):
    sheet_users = user_option_values(wb, limit=50)
    if not args:
        return sheet_users[:limit]

    pc_users = pc_user_name_values(args)
    if pc_users:
        picked = [name for name in sheet_users if name in pc_users]
        if len(picked) < limit:
            for name in pc_users:
                if name not in picked:
                    picked.append(name)
                if len(picked) >= limit:
                    break
        return picked[:limit]
    return sheet_users[:limit]


def market_activity_options(args, limit=20):
    cache = getattr(args, "_market_activity_cache", None)
    if cache is None:
        status, data = request_json(
            "GET",
            f"{pc_url(args.api)}/api/pc/market_activities/simple_market_activities",
            token=args.token,
            params={"page": 1, "per_page": limit},
            timeout=30,
        )
        if status == 200:
            cache = (data.get("data") or {}).get("list") or []
        else:
            cache = []
        args._market_activity_cache = cache
    return cache or []


def parent_customer_options(args, limit=20):
    cache = getattr(args, "_parent_customer_cache", None)
    if cache is None:
        try:
            status, data = request_json(
                "GET",
                f"{pc_url(args.api)}/api/pc/customers/selector",
                token=args.token,
                params={"page": 1, "per_page": limit, "without_count": "true"},
                timeout=30,
            )
            items = (data.get("data") or {}).get("list") or [] if status == 200 else []
            cache = [
                item
                for item in items
                if item.get("name")
                and len([part for part in str(item.get("path") or "").split("/") if part]) <= 3
            ]
        except requests.RequestException:
            cache = []
        args._parent_customer_cache = cache
    return cache or []


def random_data_relation_value(args):
    form_id = (
        getattr(args, "data_relation_custom_form_id", None)
        or os.environ.get("WWJ_DATA_RELATION_CUSTOM_FORM_ID")
        or DEFAULT_DATA_RELATION_CUSTOM_FORM_ID
    )
    cache = getattr(args, "_data_relation_cache", None)
    if cache and cache.get("custom_form_id") == form_id:
        models = cache.get("models") or []
    else:
        status, data = request_json(
            "GET",
            f"{pc_url(args.api)}/apaas/api/v2/form_entities/simple",
            token=args.token,
            params={"page": 1, "per_page": 20, "without_count": "true", "custom_form_id": form_id},
            timeout=30,
        )
        models = (data.get("data") or {}).get("models") or [] if status == 200 else []
        args._data_relation_cache = {"custom_form_id": form_id, "models": models}

    if not models:
        return None

    item = random.choice(models)
    value = item.get("label")
    if value in (None, ""):
        value = item.get("value")
    if value in (None, ""):
        value = item.get("id")
    return str(value) if value not in (None, "") else None


def is_related_product_field(name):
    return "关联产品" in name


def related_product_standard_price(context):
    product = context.get("product") or {}
    raw = product.get("raw") or {}
    value = raw.get("standard_unit_price")
    if value in (None, ""):
        value = raw.get("sale_price") or raw.get("price") or raw.get("unit_price")
    try:
        return float(value)
    except (TypeError, ValueError):
        return 100


def related_product_quantity(context):
    return 1


def related_product_total_amount(context):
    return round(related_product_standard_price(context) * related_product_quantity(context), 2)


def template_example_value(ws, col):
    value = ws.cell(8, col).value
    if isinstance(value, str) and value.startswith("("):
        return None
    if value == 0:
        return None
    return value if value not in (None, "") else None


def random_token(length=8):
    return uuid.uuid4().hex[:length]


def resolve_rows(args):
    if getattr(args, "rows_random", None):
        low, high = args.rows_random
        if low > high:
            low, high = high, low
        args.rows = random.randint(low, high)
    if not getattr(args, "rows", None):
        args.rows = 1


def full_field_value(wb, ws, header, col, suffix, index, context=None, module_key=None, args=None):
    name = clean_header(header)
    lower_name = name.lower()
    context = context or {}
    if name.startswith("CRM_期初应收款余额"):
        return None
    if name == "单行成本":
        return 10
    if name == "A小数1默认10":
        return 0.01
    if name == "额外优惠金额-30~2百":
        return 10
    if name == "金额规范化的角色VB发送":
        return 100
    if name == "开票金额-50~300":
        return 100
    if name == "金额地方":
        return 100
    if "上级客户" in name:
        if "ID" in name:
            return None
        return (context.get("parent_customer") or {}).get("name")
    if module_key == "contract" and "合同标题" in name:
        return f"自动化合同{suffix}-{index}-{random_token(4)}"
    if name == "新增小数-0.1~99.":
        return 44.1
    if name == "测试删除字段提醒":
        return f"{datetime.now().strftime('%H%M%S')}{index:03d}{random.randint(100, 999)}"
    if name == "数字（单行+多行）":
        return index
    if name == "新增数字" or ("数字" in name and "公式" not in name and "计算" not in name and "单行+多行" not in name):
        return index
    if name == "小数-0.9~-0.1":
        return -0.5
    if name == "百分比整数限制":
        return 0.01
    if name == "百分比小数限制":
        return 0.015
    if name == "百分比整数":
        return 0.01
    if name == "百分比小数":
        return 0.0125
    if name == "百分比整数默认155%":
        return 1.55
    if name == "百分比小数默认-0.83":
        return -0.0083
    if name == "百分比小数默认0%":
        return 0.0
    if "对应市场活动" in name or "市场活动" in name:
        if module_key in ("lead", "lead-pool"):
            market_activity = context.get("market_activity") or {}
            if "ID" in name:
                return None
            return market_activity.get("name")
        return None
    if "店铺员工" in name:
        return None
    if "店铺" in name:
        return None
    if "咨询产品" in name:
        return None
    if name == "跟进状态":
        options = option_values_from_sheet(wb, header, limit=10, example=ws.cell(8, col).value)
        if options:
            return min(options, key=lambda item: len(str(item)))
    if "级联" in name:
        options = option_values_from_sheet(wb, header, limit=10, example=ws.cell(8, col).value)
        if options:
            return min(options, key=lambda item: len(str(item)))
        return None
    if name in ("公司类型", "经营状态", "企业类型"):
        options = option_values_from_sheet(wb, header, limit=2, example=ws.cell(8, col).value)
        return options[0] if options else None
    if ("下拉" in name or "来源" in name) and "[关联产品]单选下拉" not in name:
        options = option_values_from_sheet(wb, header, limit=2, example=ws.cell(8, col).value)
        if options:
            if "多选" in name:
                return "，".join(str(v) for v in options)
            return options[0]
    if "状态" == name:
        return None
    if name == "折扣":
        return 1
    if module_key == "product" and "产品分类" in name:
        category_header = name.replace("_ID", "").replace("ID", "")
        options = option_values_from_sheet(wb, category_header, limit=1, example=ws.cell(8, col).value)
        return options[0] if options else None
    if name == "ID" and module_key == "customer":
        return None
    if name == "ID" or "唯一性ID" in name:
        return None
    if module_key == "customer" and "客户名称" in name:
        return f"自动化客户{suffix}-{index}-{random_token(4)}"
    if "对应客户" in name and "ID" in name:
        return (context.get("customer") or {}).get("id")
    if name.startswith("对应客户"):
        return (context.get("customer") or {}).get("name")
    if name == "客户名称":
        return (context.get("customer") or {}).get("name") or f"自动化客户{suffix}-{index}"
    if "对应商机" in name and "ID" in name:
        return (context.get("opportunity") or {}).get("id")
    if name.startswith("对应商机"):
        return (context.get("opportunity") or {}).get("name")
    if name == "商机标题" and module_key != "opportunity":
        return (context.get("opportunity") or {}).get("name")
    if "对应报价单" in name and "ID" in name:
        return (context.get("quotation") or {}).get("id")
    if name.startswith("对应报价单"):
        return (context.get("quotation") or {}).get("name")
    if name == "报价单名称" and module_key != "quotation":
        return (context.get("quotation") or {}).get("name")
    if name.startswith("合同标题") and context.get("contract") and module_key != "contract":
        return context["contract"].get("name")
    if ("对应合同" in name or "合同标题" in name) and "ID" in name:
        return (context.get("contract") or {}).get("id")
    if "对应联系人" in name:
        return (context.get("contact") or {}).get("name")
    if is_related_product_field(name) and ("金额计算" in name or "总价" in name or "合计" in name):
        return None
    if is_related_product_field(name) and "折扣" in name:
        return 1
    if is_related_product_field(name) and "ID" in name:
        return None
    if is_related_product_field(name) and "产品编号" in name:
        return None
    if is_related_product_field(name) and "产品名称" in name:
        return (context.get("product") or {}).get("name")
    if is_related_product_field(name) and "售价" in name:
        return related_product_standard_price(context)
    if is_related_product_field(name) and "数量" in name:
        return related_product_quantity(context)
    if is_related_product_field(name) and "计算数字" in name:
        return 1
    if is_related_product_field(name) and "单选计算" in name:
        option = first_option_from_sheet(wb, header)
        return option
    if is_related_product_field(name) and "单选下拉" in name:
        options = option_values_from_sheet(wb, header, limit=2, example=ws.cell(8, col).value)
        if not options:
            options = option_values_from_sheet(wb, "单选下拉列表", limit=2, example=ws.cell(8, col).value)
        if not options:
            options = option_values_from_sheet(wb, "单选下拉列表1", limit=2, example=ws.cell(8, col).value)
        if not options:
            options = option_values_from_sheet(wb, "单选下拉", limit=2, example=ws.cell(8, col).value)
        return options[0] if options else None
    if is_related_product_field(name) and "备注" in name:
        return f"关联产品备注{suffix}-{index}"
    if name == "数据关联-hsu":
        return 1234
    if "数据关联" in name:
        return None
    if name == "测试公式-数字":
        return 1
    if name == "aaaa数字1":
        return 1
    if name == "aaa数字2":
        return 2
    if name == "aaa公式1":
        return 3
    if name == "xz自定义数字":
        return 1
    if name == "公式数字":
        return 1
    if name == "新增用户":
        users = selected_user_values(wb, args=args, limit=1)
        return users[0] if users else None
    if name == "AVG百分小1,百分小2":
        return 0.1
    if name == "MAX百分小1,百分小2":
        return 0.2
    if name == "SUM":
        return 0.3
    if name in ("负责人", "经手人") or "单选用户" in name or "用户单选" in name:
        users = selected_user_values(wb, args=args, limit=1)
        return users[0] if users else None
    if module_key == "customer" and "协作人" in name:
        users = selected_user_values(wb, args=args, limit=1)
        return users[0] if users else None
    if "协作人" in name or "多选用户" in name or "用户多选" in name:
        users = selected_user_values(wb, args=args, limit=2)
        return "，".join(str(v) for v in users) if users else None
    if "部门" in name:
        return None

    options = option_values_from_sheet(wb, header, limit=2, example=ws.cell(8, col).value)
    if options:
        if "多选" in name or "用户多选" in name:
            return "，".join(str(v) for v in options)
        return options[0]

    if "客户名称" in name:
        return f"自动化客户{suffix}-{index}-{random_token(4)}"
    if "产品名称" in name:
        return f"自动化产品{suffix}-{index}-{random_token(4)}"
    if "产品编号" in name and "[关联产品" not in name:
        return f"PRD{suffix}{index:03d}{random_token(4)}"
    if "商机标题" in name or name == "商机名称":
        return f"自动化商机{suffix}-{index}-{random_token(4)}"
    if "报价单名称" in name or "报价单主标题" in name:
        return f"自动化报价单{suffix}-{index}-{random_token(4)}"
    if "报价单编号" in name:
        return f"QT{suffix}{index:03d}{random_token(4)}"
    if "合同标题" in name:
        return f"自动化合同{suffix}-{index}-{random_token(4)}"
    if "合同编号" in name:
        return f"CT{suffix}{index:03d}{random_token(4)}"
    if "发票号码" in name:
        return f"INV{suffix}{index:03d}{random_token(4)}"
    if module_key == "received-payment" and "回款期次" in name:
        return None
    if "回款期次" in name:
        return index
    if name in ("产品费用合计", "所有产品总金额"):
        return related_product_total_amount(context)
    if "整单折扣" in name:
        return 1
    if "销售单位" in name or name == "单位":
        return "个"
    if "规格" in name:
        return f"规格{index}"
    if "签约人" in name:
        return f"签约人{index}"
    if "时间" in name:
        return (datetime.now() + timedelta(days=index)).strftime("%Y-%m-%d %H:%M")
    if "姓名" in name:
        return f"自动化姓名{suffix}-{index}-{random_token(4)}"
    if "公司" in name:
        return f"自动化测试公司{suffix}-{index}-{random_token(4)}"
    if name == "电话2":
        return f"14{datetime.now().strftime('%H%M%S')}{index:03d}"
    if name == "电话" or "电话" in name:
        return f"021{random.randint(60000000, 69999999)}"
    if "手机" in name:
        return f"13{datetime.now().strftime('%H%M%S')}{index:03d}"
    if "邮箱" in name:
        return f"lead{datetime.now().strftime('%H%M%S')}{index}{random_token(6)}@example.com"
    if "微信" in name:
        return f"wx{datetime.now().strftime('%H%M%S')}{index}{random_token(6)}"
    if "QQ" in name or "qq" in name:
        return f"{random.randint(100000000, 999999999)}{index}"
    if "旺旺" in name:
        return f"ww{datetime.now().strftime('%H%M%S')}{index}{random_token(4)}"
    if name == "国家":
        return "中国"
    if name == "省":
        return "上海"
    if name == "市":
        return "上海市"
    if name == "区":
        return "徐汇区"
    if "地址" in name:
        return f"上海市徐汇区自动化测试路{index}号"
    if "邮编" in name:
        return "200030"
    if "网址" in name or "链接" in name:
        return f"https://example.com/lead/{suffix}/{index}"
    if "date" in lower_name or "起始日" in name or "开始日" in name or "日期" in name:
        return (datetime.now() + timedelta(days=index)).strftime("%Y-%m-%d")
    if "生日" in name or "birth" in lower_name:
        return (datetime.now() + timedelta(days=index)).strftime("%Y-%m-%d %H:%M")
    if name == "签到时长":
        return (datetime.now() + timedelta(days=index)).strftime("%Y/%m/%d %H:%M")
    if name == "百分比":
        return 0.09
    if "百分比" in name:
        return 0.1
    if name == "自定义整数0-2":
        return 1
    if name == "回款金额-20~1000":
        return 1000
    if name == "新增金额-1~100":
        return 99.5
    if name == "新增金额":
        return 99.5
    if module_key == "opportunity" and name == "预计销售金额":
        return related_product_total_amount(context)
    if module_key == "quotation" and name in ("报价总金额", "报价单总金额"):
        return related_product_total_amount(context)
    if module_key == "contract" and (
        "合同总金额大写" in name or "合同金额（大写）" in name or "合同金额(大写)" in name
    ):
        return None
    if module_key == "contract" and "合同总金额" in name:
        return related_product_total_amount(context)
    if "销售收入" in name or "销售成本" in name:
        return round(1000 + index * 10.5, 2)
    if "整数" in name:
        return index
    if name == "自定义小数0.1~0.9":
        return 0.12
    if name == "金额22":
        return 1010
    if "小数" in name:
        return round(index + 0.25, 2)
    if "金额" in name:
        return round(1000 + index * 10.5, 2)
    if "单价" in name:
        return 100.5
    if name == "number":
        return 100
    if name == "numbertest":
        return 100
    if name == "金额22":
        return 1010
    if "currency" in lower_name:
        return 100.5
    if "计算" in name:
        return 1
    if "备注" in name:
        return f"自动化导入备注{suffix}-{index}-{random_token(4)}"
    if "多行" in name:
        return f"自动化多行文本{suffix}-{index}-{random_token(4)}"
    if "单行" in name or "文本" in name:
        return f"自动化文本{suffix}-{index}-{random_token(4)}"

    example_cell = ws.cell(8, col).value
    if isinstance(example_cell, (int, float)) and not isinstance(example_cell, bool):
        return index
    example = template_example_value(ws, col)
    return example if example is not None else f"自动化字段{suffix}-{index}"


def minimal_field_value(header, suffix, index):
    name = clean_header(header)
    if name in ("公司类型", "经营状态", "企业类型"):
        return None
    if "客户名称" in header:
        return f"自动化客户{suffix}-{index}-{random_token(4)}"
    if "产品名称" in header:
        return f"自动化产品{suffix}-{index}-{random_token(4)}"
    if "商机标题" in header or "商机名称" in header:
        return f"自动化商机{suffix}-{index}-{random_token(4)}"
    if "报价单名称" in header or "报价单主标题" in header:
        return f"自动化报价单{suffix}-{index}-{random_token(4)}"
    if "合同标题" in header:
        return f"自动化合同{suffix}-{index}-{random_token(4)}"
    if "公司" in header:
        return f"自动化测试公司{suffix}-{index}-{random_token(4)}"
    if "姓名" in header:
        return f"自动化姓名{suffix}-{index}-{random_token(4)}"
    if "手机" in header and "自定义" not in header:
        return f"139{random.randint(10000000, 99999999)}"
    if "邮箱" in header:
        return f"lead{suffix}{index}{random_token(4)}@example.com"
    return None


def generate_source_from_template(args):
    resolve_rows(args)
    src = Path(args.template)
    out = Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(src)
    ws = wb["线索XS"] if "线索XS" in wb.sheetnames else wb[wb.sheetnames[0]]
    header_row, data_start = find_template_rows(ws)
    headers = {str(ws.cell(header_row, col).value or ""): col for col in range(1, ws.max_column + 1)}

    for row in range(data_start, data_start + max(args.rows, 20)):
        for col in range(1, ws.max_column + 1):
            ws.cell(row, col).value = None

    suffix = f"{time.strftime('%H%M%S')}{uuid.uuid4().hex[:6]}"
    for i in range(1, args.rows + 1):
        row = data_start + i - 1
        for header, col in headers.items():
            if args.fill_mode == "all":
                value = full_field_value(
                    wb,
                    ws,
                    header,
                    col,
                    suffix,
                    i,
                    getattr(args, "association_context", {}),
                    getattr(args, "module_key", None),
                    args,
                )
            else:
                value = minimal_field_value(header, suffix, i)
            if args.module_key == "customer" and col == 6:
                value = None
            if value not in (None, ""):
                ws.cell(row, col).value = value

        for header, col in headers.items():
            if "必填" not in header or ws.cell(row, col).value not in (None, ""):
                continue
            option = first_option_from_sheet(wb, header)
            ws.cell(row, col).value = option if option not in (None, "") else ws.cell(8, col).value

    keep_last_row = data_start + args.rows - 1
    if ws.max_row > keep_last_row:
        ws.delete_rows(keep_last_row + 1, ws.max_row - keep_last_row)

    compact_data_validations(ws)
    wb.save(out)
    print(f"测试导入文件已生成: {out}")
    print(f"sheet={ws.title}, header_row={header_row}, data_start={data_start}, rows={args.rows}, fill_mode={args.fill_mode}")


def upload_to_oss(api, token, file_path, upload_name=None):
    user_id, org_id, user = current_user(api, token)
    status, data = request_json(
        "GET",
        f"{pc_url(api)}{upload_to_oss.qiniu_token_path}?policy=attachment",
        token=token,
        timeout=30,
    )
    if status != 200 or "uptoken" not in data:
        raise RuntimeError(f"获取 OSS token 失败: HTTP {status} {data}")

    uptoken = data["uptoken"]
    name = upload_name or Path(file_path).name
    content_type = mimetypes.guess_type(name)[0] or "application/octet-stream"
    web_origin = api.rstrip("/")
    cmd = [
        "curl",
        "-s",
        uptoken["host"],
        "-H",
        "Accept: application/json, text/plain, */*",
        "-H",
        "Accept-Language: zh-CN,zh;q=0.9",
        "-H",
        f"Origin: {web_origin}",
        "-H",
        f"Referer: {web_origin}/",
        "-H",
        "User-Agent: Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/149.0.0.0 Safari/537.36",
        "-F",
        "chunk=0",
        "-F",
        "chunks=1",
        "-F",
        f"x:userid={user_id}",
        "-F",
        f"x:orgid={org_id}",
        "-F",
        f"x:name={name}",
        "-F",
        f"x:custom_name={name}",
        "-F",
        f"policy={uptoken['policy']}",
        "-F",
        f"OSSAccessKeyId={uptoken['accessid']}",
        "-F",
        f"signature={uptoken['signature']}",
        "-F",
        f"callback={uptoken.get('callback', '')}",
        "-F",
        f"key={uuid.uuid4()}",
        "-F",
        "Content-Disposition=inline",
        "-F",
        f"file=@{file_path};filename={name};type={content_type}",
    ]
    resp = subprocess.run(cmd, check=False, capture_output=True, text=True, timeout=120)
    if resp.returncode != 0:
        raise RuntimeError(f"OSS 上传命令失败: exit={resp.returncode} stderr={resp.stderr[:1000]}")
    try:
        upload_body = json.loads(resp.stdout)
    except ValueError:
        raise RuntimeError(f"OSS 上传返回非 JSON: {resp.stdout[:1000]} {resp.stderr[:1000]}")
    payload = upload_body.get("payload", {})
    if not payload.get("id") or not payload.get("file_url"):
        raise RuntimeError(f"OSS 上传未返回附件信息: {resp.stdout[:1000]}")
    return {
        "attachment_id": payload["id"],
        "file_url": payload["file_url"],
        "user_id": user_id,
        "organization_id": org_id,
        "user_name": user.get("name"),
    }


class FayeListener:
    def __init__(self, faye_url, channel, import_client_id=None):
        self.faye_url = faye_url
        self.channel = channel
        self.import_client_id = import_client_id
        self.file_hashes = set()
        self.session = requests.Session()
        self.messages = []
        self._stop = False
        self._thread = None

    def _matches_current_import(self, msg):
        if not self.import_client_id:
            return True
        data = msg.get("data") or {}
        file_hash = data.get("file_hash")
        if data.get("import_client_id") == self.import_client_id:
            if file_hash:
                self.file_hashes.add(file_hash)
            return True
        return bool(file_hash and file_hash in self.file_hashes)

    def start(self):
        hs = self.session.post(
            self.faye_url,
            json=[{"channel": "/meta/handshake", "version": "1.0", "supportedConnectionTypes": ["long-polling"], "id": "1"}],
            timeout=15,
        ).json()[0]
        if not hs.get("successful"):
            raise RuntimeError(f"faye handshake 失败: {hs}")
        self.client_id = hs["clientId"]
        sub = self.session.post(
            self.faye_url,
            json=[{"channel": "/meta/subscribe", "clientId": self.client_id, "subscription": self.channel, "id": "2"}],
            timeout=15,
        ).json()[0]
        if not sub.get("successful"):
            raise RuntimeError(f"faye subscribe 失败: {sub}")

        def loop():
            msg_id = 3
            while not self._stop:
                try:
                    arr = self.session.post(
                        self.faye_url,
                        json=[
                            {
                                "channel": "/meta/connect",
                                "clientId": self.client_id,
                                "connectionType": "long-polling",
                                "id": str(msg_id),
                            }
                        ],
                        timeout=50,
                    ).json()
                    msg_id += 1
                    for msg in arr:
                        if msg.get("channel") == self.channel and self._matches_current_import(msg):
                            self.messages.append(msg)
                            print_json("faye_message", msg)
                except Exception:
                    pass

        self._thread = threading.Thread(target=loop, daemon=True)
        self._thread.start()

    def wait_for(self, process_type, seconds):
        deadline = time.time() + seconds
        while time.time() < deadline:
            for msg in self.messages:
                if (msg.get("data") or {}).get("process_type") == process_type:
                    return msg
            time.sleep(1)
        return None

    def stop(self):
        self._stop = True


def import_flow(args):
    resolve_business_type(args)
    file_path = Path(args.file)
    if not file_path.exists():
        raise FileNotFoundError(file_path)

    status, new_data = request_json(
        "GET",
        entity_loader_url(args, "new"),
        token=args.token,
        params={
            **{"loader_name": args.loader_name},
            **({"custom_field_template_id": args.custom_field_template_id} if getattr(args, "custom_field_template_id", "") else {}),
            **({"common_id": args.common_id} if getattr(args, "common_id", "") and args.module_key in ("lead-pool", "customer-pool") else {}),
        },
        timeout=30,
    )
    new_body = new_data.get("data") or {}
    print_json(
        "entity_loaders/new",
        {
            "http_status": status,
            "code": new_data.get("code"),
            "import_client_id": new_body.get("import_client_id"),
            "import_max_count": new_body.get("import_max_count"),
            "faye_channel": new_body.get("faye_channel"),
            "duplicate_fields": new_body.get("duplicate_fields", []),
            "recognize_fields_count": len(new_body.get("recognize_fields", [])),
            "custom_field_template_id": getattr(args, "custom_field_template_id", ""),
        },
    )
    client_id = (new_data.get("data") or {}).get("import_client_id")
    channel = (new_data.get("data") or {}).get("faye_channel")
    if not client_id:
        raise RuntimeError("new 接口未返回 import_client_id")

    listener = None
    if args.listen_faye and channel:
        listener = FayeListener(args.faye_url, channel, client_id)
        listener.start()
        print(f"已监听 faye: {channel}")

    upload_to_oss.qiniu_token_path = args.qiniu_token_path
    upload = upload_to_oss(args.api, args.token, file_path, getattr(args, "upload_name", None))
    local_stat = {"size": file_path.stat().st_size, "sha1": file_sha1(file_path)}
    try:
        remote_stat = remote_file_digest(upload["file_url"])
    except Exception as exc:
        remote_stat = {"error": str(exc)}
    upload["local_file"] = str(file_path)
    upload["local_stat"] = local_stat
    upload["remote_stat"] = remote_stat
    print_json("oss_upload", upload)

    upload_payload = {
        "loader_name": args.loader_name,
        "import_client_id": client_id,
        "import_max_count": new_body.get("import_max_count") or 10000,
        "import_trigger_approve_max_count": new_body.get("import_trigger_approve_max_count") or 1000,
        "recognize_fields": new_body.get("recognize_fields", []),
        "entity_loader_file": upload["file_url"],
        "file_attachment_id": upload["attachment_id"],
    }
    if getattr(args, "custom_field_template_id", ""):
        upload_payload["custom_field_template_id"] = args.custom_field_template_id
    if getattr(args, "common_id", "") and args.module_key in ("lead-pool", "customer-pool"):
        upload_payload["common_id"] = args.common_id
    status, upload_data = request_json(
        "POST", entity_loader_url(args, "upload"), token=args.token, json=upload_payload, timeout=60
    )
    print_json("entity_loaders/upload", {"http_status": status, "body": upload_data})
    if status != 200 or str(upload_data.get("code")) != "0":
        print_json(
            "upload_failed_reproduce",
            {
                "loader_name": args.loader_name,
                "import_client_id": client_id,
                "entity_loader_file": upload["file_url"],
                "file_attachment_id": upload["attachment_id"],
                "local_file": str(file_path),
                "local_stat": local_stat,
                "remote_stat": remote_stat,
            },
        )
        if "正在进行的非系统模板导入任务" in str(upload_data.get("message", "")):
            raise RuntimeError(
                "上传导入文件失败: 当前账号已有未完成的非系统模板上传/校验任务，"
                "请先在导入弹窗删除已上传文件或完成/取消该任务后重试"
            )
        raise RuntimeError(f"上传导入文件失败: HTTP {status} {upload_data}")

    validated = listener.wait_for("validated_result", args.validate_wait) if listener else None
    if listener and not validated:
        print(f"等待 {args.validate_wait}s 后未收到 validated_result")

    import_params = {
        "loader_name": args.loader_name,
        "entity_loader_file": upload["file_url"],
        "file_attachment_id": upload["attachment_id"],
        "import_client_id": client_id,
    }
    if getattr(args, "custom_field_template_id", ""):
        import_params["custom_field_template_id"] = args.custom_field_template_id
    if getattr(args, "common_id", "") and args.module_key in ("lead-pool", "customer-pool"):
        import_params["common_id"] = args.common_id
    status, import_data = request_json(
        "POST", entity_loader_url(args, "import"), token=args.token, json=import_params, timeout=60
    )
    print_json("entity_loaders/import", {"http_status": status, "body": import_data})

    imported = listener.wait_for("imported_result", args.import_wait) if listener else None
    import_result = {
        "loader_name": args.loader_name,
        "attachment_id": upload["attachment_id"],
        "status": "unknown",
        "total_count": None,
        "success_count": None,
        "fail_count": None,
        "error_file": None,
        "result_display": None,
    }
    if imported:
        imported_data = imported.get("data") or {}
        import_result.update(
            {
                "status": "success" if not imported_data.get("has_error_data") else "failed",
                "total_count": imported_data.get("total_count"),
                "success_count": imported_data.get("total_count") if not imported_data.get("has_error_data") else 0,
                "fail_count": (imported_data.get("error_info") or {}).get("fail_count"),
                "error_file": (imported_data.get("error_info") or {}).get("file_path"),
                "result_display": imported_data.get("result_display"),
            }
        )
    if listener and not imported:
        print(f"等待 {args.import_wait}s 后未收到 imported_result")
        import_result["status"] = "timeout"
    if listener:
        listener.stop()

    history = query_history(args)
    latest = (((history.get("body") or {}).get("data") or {}).get("list") or [{}])[0] if history else {}
    if latest.get("attachment_id") == upload["attachment_id"]:
        import_result.update(
            {
                "status": "success" if latest.get("fail_count") == 0 and latest.get("success_count", 0) > 0 else "failed",
                "total_count": latest.get("total_count"),
                "success_count": latest.get("success_count"),
                "fail_count": latest.get("fail_count"),
                "error_file": latest.get("file_path"),
                "result_display": latest.get("result_display"),
            }
        )
    query_latest_leads(args)
    print_json("import_result_summary", import_result)
    return import_result


def query_history(args):
    status, data = request_json(
        "GET",
        entity_loader_url(args, "import_histories"),
        token=args.token,
        params={"importable_type": args.loader_name, "page": 1, "per_page": args.per_page},
        timeout=30,
    )
    result = {"http_status": status, "body": data}
    print_json("import_histories", result)
    return result


def query_latest_leads(args):
    module_config = MODULE_CONFIGS.get(args.module_key, MODULE_CONFIGS["lead"])
    api_path = module_config["api_path"]
    params = {"per_page": args.per_page, "sort": "created_at", "order": "desc", "page": 1}
    if args.module_key == "lead-pool" and getattr(args, "common_id", ""):
        params["lead_common_setting_id"] = args.common_id
    if args.module_key == "customer-pool" and getattr(args, "common_id", ""):
        params["customer_common_setting_id"] = args.common_id
    status, data = request_json(
        "GET",
        f"{args.api.rstrip('/')}/api/v2/{api_path}",
        token=args.token,
        params=params,
        timeout=30,
    )
    body = data.get("data") or {} if isinstance(data, dict) else {}
    items = body.get(api_path) or body.get("list") or []
    summary = [
        {
            "id": item.get("id"),
            "name": item.get("name") or item.get("title") or item.get("sn") or item.get("product_name"),
            "company_name": item.get("company_name") or item.get("customer_name"),
            "created_at": item.get("created_at"),
        }
        for item in items
    ]
    print_json(f"latest_{api_path}", {"http_status": status, "items": summary})


def related_item_context(item):
    customer = item.get("customer") if isinstance(item.get("customer"), dict) else {}
    opportunity = item.get("opportunity") if isinstance(item.get("opportunity"), dict) else {}
    return {
        "id": item.get("id"),
        "name": item.get("name") or item.get("title") or item.get("contract_name") or item.get("product_name"),
        "no": item.get("product_no") or item.get("no") or item.get("sn"),
        "customer_id": item.get("customer_id") or customer.get("id"),
        "customer_name": item.get("customer_name") or item.get("customer.name") or customer.get("name"),
        "opportunity_id": item.get("opportunity_id") or opportunity.get("id"),
        "opportunity_name": item.get("opportunity_name") or item.get("opportunity.name") or opportunity.get("title"),
        "raw": item,
    }


def first_item(args, module_key):
    config = MODULE_CONFIGS[module_key]
    path = config["api_path"]
    items = []
    for attempt in range(3):
        status, data = request_json(
            "GET",
            f"{args.api.rstrip('/')}/api/v2/{path}",
            token=args.token,
            params={"per_page": 1, "sort": "created_at", "order": "desc", "page": 1},
            timeout=30,
        )
        body = data.get("data") or {} if isinstance(data, dict) else {}
        items = body.get(path) or body.get("list") or []
        if status == 200 and items:
            break
        time.sleep(attempt + 1)
    if not items:
        return {}
    item = items[0]
    return related_item_context(item)


def first_item_with_value(args, module_key, field_name, pages=5):
    config = MODULE_CONFIGS[module_key]
    path = config["api_path"]
    candidates = []
    for page in range(1, pages + 1):
        status, data = request_json(
            "GET",
            f"{args.api.rstrip('/')}/api/v2/{path}",
            token=args.token,
            params={"per_page": 50, "sort": "created_at", "order": "desc", "page": page},
            timeout=30,
        )
        body = data.get("data") or {} if isinstance(data, dict) else {}
        items = body.get(path) or body.get("list") or []
        for item in items:
            value = item.get(field_name)
            if value not in (None, ""):
                candidates.append(related_item_context(item))
    return random.choice(candidates) if candidates else {}


def first_item_for_customer(args, module_key, customer_id, fallback=True):
    if not customer_id:
        return first_item(args, module_key)
    config = MODULE_CONFIGS[module_key]
    path = config["api_path"]
    for page in range(1, 6):
        try:
            status, data = request_json(
                "GET",
                f"{args.api.rstrip('/')}/api/v2/{path}",
                token=args.token,
                params={"per_page": 50, "sort": "created_at", "order": "desc", "page": page},
                timeout=30,
            )
        except requests.RequestException:
            continue
        body = data.get("data") or {} if isinstance(data, dict) else {}
        items = body.get(path) or body.get("list") or []
        for item in items:
            if str(item.get("customer_id") or "") == str(customer_id):
                return related_item_context(item)
    if fallback:
        try:
            return first_item(args, module_key)
        except requests.RequestException:
            return {}
    return {}


def strict_first_item_for_customer(args, module_key, customer_id, pages=5):
    if not customer_id:
        return {}
    config = MODULE_CONFIGS[module_key]
    path = config["api_path"]
    for page in range(1, pages + 1):
        try:
            status, data = request_json(
                "GET",
                f"{args.api.rstrip('/')}/api/v2/{path}",
                token=args.token,
                params={"per_page": 50, "sort": "created_at", "order": "desc", "page": page},
                timeout=30,
            )
        except requests.RequestException:
            return {}
        body = data.get("data") or {} if isinstance(data, dict) else {}
        items = body.get(path) or body.get("list") or []
        for item in items:
            if str(item.get("customer_id") or "") == str(customer_id):
                return related_item_context(item)
    return {}


def list_related_items(args, module_key, pages=3, per_page=50):
    config = MODULE_CONFIGS[module_key]
    path = config["api_path"]
    results = []
    for page in range(1, pages + 1):
        try:
            status, data = request_json(
                "GET",
                f"{args.api.rstrip('/')}/api/v2/{path}",
                token=args.token,
                params={"per_page": per_page, "sort": "created_at", "order": "desc", "page": page},
                timeout=30,
            )
        except requests.RequestException:
            continue
        body = data.get("data") or {} if isinstance(data, dict) else {}
        items = body.get(path) or body.get("list") or []
        results.extend(related_item_context(item) for item in items)
    return results


def choose_customer_candidate(candidates, avoid_customer_ids=None):
    avoid_customer_ids = {str(item) for item in (avoid_customer_ids or []) if item not in (None, "")}
    unique = []
    seen = set()
    for candidate in candidates:
        customer_id = candidate.get("customer", {}).get("id")
        if not customer_id or str(customer_id) in seen:
            continue
        seen.add(str(customer_id))
        unique.append(candidate)
    if not unique:
        return {}
    preferred = [item for item in unique if str(item.get("customer", {}).get("id")) not in avoid_customer_ids]
    return random.choice(preferred or unique)


def choose_customer_anchor(items, avoid_customer_ids=None):
    avoid_customer_ids = {str(item) for item in (avoid_customer_ids or []) if item not in (None, "")}
    unique = []
    seen = set()
    for item in items:
        customer_id = item.get("customer_id")
        if not customer_id or str(customer_id) in seen:
            continue
        seen.add(str(customer_id))
        unique.append(item)
    if not unique:
        return {}
    preferred = [item for item in unique if str(item.get("customer_id")) not in avoid_customer_ids]
    return random.choice(preferred or unique)


def visible_customer_ids(args):
    if hasattr(args, "_visible_customer_ids"):
        return args._visible_customer_ids
    ids = {
        str(item.get("id"))
        for item in list_related_items(args, "customer", pages=1, per_page=100)
        if item.get("id") not in (None, "")
    }
    args._visible_customer_ids = ids
    return ids


def find_customer_chain(args, prefer_module=None):
    opportunities = list_related_items(args, "opportunity", pages=1, per_page=100)
    random.shuffle(opportunities)
    avoid_customer_ids = getattr(args, "avoid_customer_ids", None)
    preferred = [item for item in opportunities if str(item.get("customer_id")) not in {str(i) for i in (avoid_customer_ids or [])}]
    visible_ids = visible_customer_ids(args)
    for opportunity in preferred + [item for item in opportunities if item not in preferred]:
        customer_id = opportunity.get("customer_id")
        if not customer_id or (visible_ids and str(customer_id) not in visible_ids):
            continue
        contact = strict_first_item_for_customer(args, "contact", customer_id, pages=1)
        quotation = strict_first_item_for_customer(args, "quotation", customer_id, pages=1)
        contract = strict_first_item_for_customer(args, "contract", customer_id, pages=1)
        if prefer_module == "quotation" and not quotation:
            continue
        return {
            "customer": {
                "id": customer_id,
                "name": opportunity.get("customer_name"),
            },
            "contact": contact,
            "opportunity": opportunity,
            "quotation": quotation,
            "contract": contract,
        }
    return {}


def find_contract_chain(args):
    visible_ids = visible_customer_ids(args)
    contracts = [
        item for item in list_related_items(args, "contract", pages=1, per_page=100)
        if not visible_ids or str(item.get("customer_id")) in visible_ids
    ]
    contract = choose_customer_anchor(
        contracts,
        avoid_customer_ids=getattr(args, "avoid_customer_ids", None),
    )
    customer = customer_context_from_related_item(contract)
    customer_id = customer.get("id")
    if not customer_id:
        return {}
    return {
        "customer": customer,
        "contact": strict_first_item_for_customer(args, "contact", customer_id, pages=1),
        "opportunity": strict_first_item_for_customer(args, "opportunity", customer_id, pages=1),
        "quotation": strict_first_item_for_customer(args, "quotation", customer_id, pages=1),
        "contract": contract,
    }


def customer_context_from_related_item(item):
    customer_id = item.get("customer_id")
    customer_name = item.get("customer_name") or (item.get("raw") or {}).get("customer.name")
    if not customer_id:
        return {}
    return {
        "id": customer_id,
        "name": customer_name,
    }


def build_association_context(args):
    def safe_first_item(module_key):
        try:
            return first_item(args, module_key)
        except Exception as exc:
            print(f"{module_key} 关联数据查询失败，继续使用空值: {exc}")
            return {}

    if args.module_key in ("lead", "lead-pool"):
        activities = market_activity_options(args, limit=100)
        context = {"market_activity": random.choice(activities) if activities else {}}
        print_json("association_context", context)
        return context

    if args.module_key == "customer":
        parents = parent_customer_options(args)
        context = {"parent_customer": random.choice(parents) if parents else {}}
        print_json(
            "association_context",
            {"parent_customer": {key: context["parent_customer"].get(key) for key in ("id", "name", "path")}},
        )
        return context

    if args.module_key == "contact":
        customer = safe_first_item("customer")
        context = {
            "customer": customer,
            "product": first_item_with_value(args, "product", "product_no", pages=1) or safe_first_item("product"),
        }
        print_json(
            "association_context",
            {key: {k: v for k, v in value.items() if k != "raw"} for key, value in context.items()},
        )
        return context

    if args.module_key in ("payment-plan", "received-payment", "invoiced-payment"):
        chained = find_contract_chain(args)
        contract = chained.get("contract", {})
        if not contract.get("id"):
            raise RuntimeError("未找到可用于导入的合同，回款/开票模块需要先存在一条合同数据")
        context = {
            "customer": chained.get("customer", {}),
            "contact": chained.get("contact", {}),
            "opportunity": chained.get("opportunity", {}),
            "quotation": chained.get("quotation", {}),
            "contract": contract,
            "product": first_item_with_value(args, "product", "product_no", pages=1) or safe_first_item("product"),
        }
        print_json(
            "association_context",
            {key: {k: v for k, v in value.items() if k != "raw"} for key, value in context.items()},
        )
        return context

    chained = find_customer_chain(args, prefer_module="quotation" if args.module_key == "contract" else None) if args.module_key in ("quotation", "contract") else {}
    if chained:
        context = {
            "customer": chained["customer"],
            "contact": chained["contact"],
            "opportunity": chained["opportunity"],
            "quotation": chained["quotation"],
            "contract": chained["contract"],
            "product": first_item_with_value(args, "product", "product_no", pages=1) or safe_first_item("product"),
        }
        print_json(
            "association_context",
            {key: {k: v for k, v in value.items() if k != "raw"} for key, value in context.items()},
        )
        return context

    customer = safe_first_item("customer")
    customer_id = customer.get("id")
    opportunity = first_item_for_customer(args, "opportunity", customer_id, fallback=False) if customer_id else {}
    if not opportunity and args.module_key in ("quotation", "contract"):
        opportunity = safe_first_item("opportunity")
        customer = customer_context_from_related_item(opportunity) or customer
        customer_id = customer.get("id")
    context = {
        "customer": customer,
        "contact": first_item_for_customer(args, "contact", customer_id, fallback=False) if customer_id else safe_first_item("contact"),
        "opportunity": opportunity or (first_item_for_customer(args, "opportunity", customer_id, fallback=False) if customer_id else safe_first_item("opportunity")),
        "quotation": first_item_for_customer(args, "quotation", customer_id, fallback=False) if customer_id else safe_first_item("quotation"),
        "contract": first_item_for_customer(args, "contract", customer_id, fallback=False) if customer_id else safe_first_item("contract"),
        "product": first_item_with_value(args, "product", "product_no", pages=1) or safe_first_item("product"),
    }
    if args.module_key in ("payment-plan", "received-payment", "invoiced-payment"):
        contract = context.get("contract") or {}
        if not contract.get("id"):
            raise RuntimeError("未找到可用于导入的合同，回款/开票模块需要先存在一条合同数据")
        if contract.get("customer_id"):
            context["customer"] = {
                "id": contract.get("customer_id"),
                "name": contract.get("customer_name"),
            }
    print_json(
        "association_context",
        {key: {k: v for k, v in value.items() if k != "raw"} for key, value in context.items()},
    )
    return context


def resolve_config(args):
    env_name = getattr(args, "env", None) or os.environ.get("WWJ_IMPORT_ENV") or "test"
    if env_name not in ENV_CONFIGS:
        raise SystemExit(f"未知环境: {env_name}，可选: {', '.join(ENV_CONFIGS)}")

    config = ENV_CONFIGS[env_name]
    profile = load_config(env_name, getattr(args, "profile", None))
    args.env = env_name
    args.module_key = normalize_module(getattr(args, "module", None))
    args.module_config = MODULE_CONFIGS[args.module_key]
    if hasattr(args, "loader_name") and not args.loader_name:
        args.loader_name = args.module_config["loader_name"]
    args.api = (
        getattr(args, "api", None)
        or os.environ.get("WWJ_IMPORT_API")
        or profile.get("api")
        or config["api"]
    )
    args.token = getattr(args, "token", None) or os.environ.get("WWJ_USER_TOKEN") or profile.get("token")
    if args.token and args.token.startswith("user_token="):
        args.token = args.token.split("=", 1)[1]
    if hasattr(args, "entity_loader_prefix"):
        args.entity_loader_prefix = (
            args.entity_loader_prefix
            or os.environ.get("WWJ_ENTITY_LOADER_PREFIX")
            or config["entity_loader_prefix"]
        )
    if hasattr(args, "entity_loader_new_prefix"):
        args.entity_loader_new_prefix = (
            args.entity_loader_new_prefix
            or os.environ.get("WWJ_ENTITY_LOADER_NEW_PREFIX")
            or config["entity_loader_new_prefix"]
        )
    if hasattr(args, "entity_loader_history_prefix"):
        args.entity_loader_history_prefix = (
            args.entity_loader_history_prefix
            or os.environ.get("WWJ_ENTITY_LOADER_HISTORY_PREFIX")
            or config["entity_loader_history_prefix"]
        )
    if hasattr(args, "qiniu_token_path"):
        args.qiniu_token_path = (
            args.qiniu_token_path
            or os.environ.get("WWJ_QINIU_TOKEN_PATH")
            or config["qiniu_token_path"]
        )
    if hasattr(args, "faye_url"):
        args.faye_url = args.faye_url or os.environ.get("WWJ_FAYE_URL") or config["faye_url"]
    if hasattr(args, "business_type_id"):
        args.business_type_id = args.business_type_id or os.environ.get("WWJ_BUSINESS_TYPE_ID")
    if hasattr(args, "business_type_name"):
        args.business_type_name = args.business_type_name or os.environ.get("WWJ_BUSINESS_TYPE_NAME")
    if hasattr(args, "business_type_map"):
        args.business_type_map = args.business_type_map or os.environ.get("WWJ_BUSINESS_TYPE_MAP")
    if hasattr(args, "common_id"):
        args.common_id = args.common_id or os.environ.get("WWJ_COMMON_ID")

    if args.command != "generate-source-from-template" and not args.token:
        raise SystemExit(
            f"缺少 token：请在config.{env_name}.json中配置，或传 --token，"
            "或设置环境变量 WWJ_USER_TOKEN"
        )

    if args.command != "generate-source-from-template":
        print(f"使用环境: {args.env}, api={args.api}")


def full_flow(args):
    resolve_rows(args)
    out_dir = Path(args.output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    module_name = args.module_key
    template = out_dir / f"{module_name}-template.xlsx"
    source = out_dir / f"{module_name}-source.xlsx"

    browser_template = matching_local_business_template(args)
    args.output = str(template)
    download_template(args)
    browser_template = browser_template or local_browser_template(args, template)
    if browser_template and browser_template.exists():
        shutil.copyfile(browser_template, template)
        print(f"使用页面下载模板包: {browser_template}")
        if not getattr(args, "upload_name", None):
            args.upload_name = browser_template.name
    elif not has_preservable_template_package(template):
        raise RuntimeError(
            f"{args.module_config['label']} 缺少可保留包结构的系统模板。"
            "请从测试环境页面下载该模块导入模板，放到 --template-dir 指定目录，"
            "或用 --template-file 显式指定。"
        )

    args.template = str(template)
    args.output = str(source)
    args.association_context = (
        build_association_context(args)
        if args.fill_mode == "all" and args.module_key in ASSOCIATION_CONTEXT_MODULES
        else {}
    )
    if has_preservable_template_package(template):
        generate_source_preserving_package(args)
    else:
        generate_source_from_template(args)

    args.file = str(source)
    return import_flow(args)


def patch_xlsx_first_sheet_cells(path, updates):
    xlsx_path = Path(path)
    with ZipFile(xlsx_path) as zin:
        data = {name: zin.read(name) for name in zin.namelist()}

    shared_root, shared_items, strings = read_shared_strings(data)
    shared_strings_created = False
    if shared_root is None:
        shared_root = ET.Element(sheet_q("sst"), {"count": "0", "uniqueCount": "0"})
        shared_items = []
        strings = []
        shared_strings_created = True

    sheet = ET.fromstring(data["xl/worksheets/sheet1.xml"])
    sheet_data = sheet.find(sheet_q("sheetData"))
    rows_by_num = {int(row.get("r")): row for row in sheet_data.findall(sheet_q("row")) if row.get("r")}
    for row_num, col_values in updates.items():
        row_el = rows_by_num.get(row_num)
        if row_el is None:
            row_el = ET.SubElement(sheet_data, sheet_q("row"), {"r": str(row_num)})
            rows_by_num[row_num] = row_el
        for col, value in col_values.items():
            cell_ref = f"{get_column_letter(col)}{row_num}"
            if value in (None, ""):
                for cell in row_el.findall(sheet_q("c")):
                    if cell.get("r") == cell_ref:
                        row_el.remove(cell)
                        break
                continue
            set_sheet_cell(row_el, cell_ref, value, shared_root, shared_items, strings)
        sort_sheet_row_cells(row_el)

    shared_root.set("count", str(len(strings)))
    shared_root.set("uniqueCount", str(len(strings)))
    data["xl/sharedStrings.xml"] = ET.tostring(shared_root, encoding="utf-8", xml_declaration=True)
    ensure_package_metadata(data, shared_strings_created=shared_strings_created)
    data["xl/worksheets/sheet1.xml"] = ET.tostring(sheet, encoding="utf-8", xml_declaration=True)

    with ZipFile(xlsx_path, "w", ZIP_DEFLATED) as zout:
        for name, content in data.items():
            zout.writestr(name, content)


def invalid_field_value(header):
    name = clean_header(header)
    if not name or "ID" in name or "唯一性ID" in name:
        return None
    return "自动化错误字段"


def invalidate_all_input_fields(path):
    wb = load_workbook(path, read_only=True, data_only=False)
    ws = wb.active
    header_row, data_start = find_template_rows(ws)
    updates = {}
    columns = []
    for col in range(1, ws.max_column + 1):
        header = str(ws.cell(header_row, col).value or "")
        value = invalid_field_value(header)
        if value not in (None, ""):
            columns.append(col)

    for row in range(data_start, ws.max_row + 1):
        if any(ws.cell(row, col).value not in (None, "") for col in range(1, ws.max_column + 1)):
            updates[row] = {col: "自动化错误字段" for col in columns}
    patch_xlsx_first_sheet_cells(path, updates)
    return data_start, ws.max_row, len(columns)


def download_error_report(error_file, output_path):
    if not error_file:
        raise RuntimeError("导入失败结果中没有错误报告下载地址")
    resp = requests.get(error_file, timeout=120)
    resp.raise_for_status()
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_bytes(resp.content)
    print(f"错误模板已下载: {output} ({output.stat().st_size} bytes)")
    return output


def error_report_archive_info(path):
    with ZipFile(path) as archive:
        names = archive.namelist()
    if "[Content_Types].xml" in names and "xl/workbook.xml" in names:
        return {"type": "xlsx", "entries": len(names)}
    return {"type": "zip", "entries": names}


def repair_error_template_from_baseline(error_template, valid_source, repaired_output):
    src = Path(error_template)
    out = Path(repaired_output)
    if src != out:
        shutil.copyfile(src, out)

    # 错误报告需要按“源文件位置”随机读取基准行。read_only 工作簿的
    # cell() 会反复扫描 XML，200 行分片已足以让回填在写入前超时。
    error_wb = load_workbook(out, read_only=False, data_only=False)
    error_ws = error_wb.active
    error_header_row, error_data_start = find_template_rows(error_ws)
    baseline_wb = load_workbook(valid_source, read_only=False, data_only=False)
    baseline_ws = baseline_wb.active
    baseline_header_row, baseline_data_start = find_template_rows(baseline_ws)
    baseline_columns = {
        clean_header(baseline_ws.cell(baseline_header_row, col).value): col
        for col in range(1, baseline_ws.max_column + 1)
        if clean_header(baseline_ws.cell(baseline_header_row, col).value)
    }
    source_row_column = next(
        (
            col
            for col in range(1, error_ws.max_column + 1)
            if clean_header(error_ws.cell(error_header_row, col).value) == "源文件位置"
        ),
        None,
    )

    updates = {}
    for error_row in range(error_data_start, error_ws.max_row + 1):
        source_value = error_ws.cell(error_row, source_row_column).value if source_row_column else None
        source_match = re.search(r"第\s*(\d+)\s*行", str(source_value or ""))
        source_index = int(source_match.group(1)) if source_match else error_row - error_data_start + 1
        baseline_row = baseline_data_start + source_index - 1
        if baseline_row > baseline_ws.max_row:
            break
        row_updates = {}
        for error_col in range(1, error_ws.max_column + 1):
            header = clean_header(error_ws.cell(error_header_row, error_col).value)
            baseline_col = baseline_columns.get(header)
            if baseline_col:
                row_updates[error_col] = baseline_ws.cell(baseline_row, baseline_col).value
        if row_updates:
            updates[error_row] = row_updates
    patch_xlsx_first_sheet_cells(out, updates)
    print(f"错误模板已按基准数据全字段修复: {out}")
    return out


def error_template_flow(args):
    resolve_rows(args)
    out_dir = Path(args.output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    module_name = args.module_key
    template = out_dir / f"{module_name}-error-template-base.xlsx"
    valid_source = out_dir / f"{module_name}-error-source-valid.xlsx"
    invalid_source = out_dir / f"{module_name}-error-source-invalid.xlsx"
    error_template = out_dir / f"{module_name}-error-report.xlsx"
    repaired_source = out_dir / f"{module_name}-error-source-repaired.xlsx"

    browser_template = matching_local_business_template(args)
    args.output = str(template)
    download_template(args)
    browser_template = browser_template or local_browser_template(args, template)
    if browser_template and browser_template.exists():
        shutil.copyfile(browser_template, template)
        print(f"使用页面下载模板包: {browser_template}")
        args.upload_name = browser_template.name

    args.template = str(template)
    args.output = str(valid_source)
    args.association_context = (
        build_association_context(args)
        if args.fill_mode == "all" and args.module_key in ASSOCIATION_CONTEXT_MODULES
        else {}
    )
    generate_source_preserving_package(args)

    shutil.copyfile(valid_source, invalid_source)
    data_start, max_row, invalid_columns = invalidate_all_input_fields(invalid_source)
    print(
        f"已构造全字段错误数据: 第 {data_start}-{max_row} 行，"
        f"覆盖 {invalid_columns} 个可写字段"
    )

    original_upload_name = getattr(args, "upload_name", None)
    args.file = str(invalid_source)
    first_result = import_flow(args)
    if first_result.get("status") != "failed" or not first_result.get("error_file"):
        raise RuntimeError(f"预期首次导入失败并产生错误模板，但结果为: {first_result}")

    download_error_report(first_result["error_file"], error_template)
    if getattr(args, "verify_error_zip", False):
        archive_info = error_report_archive_info(error_template)
        print_json("error_report_archive", archive_info)
        if archive_info["type"] != "zip":
            raise RuntimeError(f"预期错误报告为 ZIP 压缩包，实际为 {archive_info['type']}")
        return {
            "valid_source": str(valid_source),
            "first_import": first_result,
            "error_template": str(error_template),
            "error_report_archive": archive_info,
        }
    repair_error_template_from_baseline(error_template, valid_source, repaired_source)

    args.file = str(repaired_source)
    args.upload_name = original_upload_name
    second_result = import_flow(args)
    if second_result.get("status") != "success":
        raise RuntimeError(f"错误模板修复后再次导入未成功: {second_result}")

    summary = {
        "valid_source": str(valid_source),
        "first_import": first_result,
        "error_template": str(error_template),
        "repaired_template": str(repaired_source),
        "second_import": second_result,
    }
    print_json("error_template_flow_summary", summary)
    return summary


def lead_error_template_flow(args):
    args.module = "lead"
    return error_template_flow(args)


def parse_modules(value):
    if not value or value == "all":
        return list(MODULE_CONFIGS)
    return [normalize_module(item.strip()) for item in value.split(",") if item.strip()]


def batch_full_flow(args):
    results = []
    for module_key in parse_modules(args.modules):
        child = argparse.Namespace(**vars(args))
        child.module = module_key
        child.module_key = module_key
        child.module_config = MODULE_CONFIGS[module_key]
        child.loader_name = child.module_config["loader_name"]
        child.output_dir = str(Path(args.output_dir) / module_key)
        print(f"\n===== 开始处理模块: {module_key} / {child.module_config['label']} =====")
        try:
            result = full_flow(child)
            results.append({"module": module_key, "label": child.module_config["label"], **(result or {})})
        except Exception as exc:
            result = {"module": module_key, "label": child.module_config["label"], "status": "failed", "error": str(exc)}
            results.append(result)
            print_json("module_failed", result)
            if not args.continue_on_error:
                raise
    print_json("batch_summary", results)


def build_parser():
    parser = argparse.ArgumentParser(description="CRM 导入完整流程测试工具")
    common = argparse.ArgumentParser(add_help=False)
    common.add_argument("--env", choices=sorted(ENV_CONFIGS), default=None, help="预置环境，默认 test")
    common.add_argument("--profile", choices=["gray", "standard"], default=None, help="同环境下的企业配置和模板目录")
    common.add_argument("--api", default=None, help="覆盖环境域名")
    common.add_argument("--token", default=None, help="user_token；也可用 WWJ_USER_TOKEN")
    common.add_argument("--module", default="lead", help="业务模块，例如 lead/lead-pool/customer/contact/opportunity")
    common.add_argument("--loader-name", default=None, help="覆盖模块对应的 loader_name")
    common.add_argument("--entity-loader-prefix", default=None, help="覆盖导入接口前缀")
    common.add_argument("--entity-loader-new-prefix", default=None, help="覆盖 new 接口前缀")
    common.add_argument("--entity-loader-history-prefix", default=None, help="覆盖导入历史接口前缀")
    common.add_argument("--qiniu-token-path", default=None, help="覆盖 OSS token 接口路径")
    common.add_argument("--business-type-id", default=None, help="客户/商机/报价单/合同导入使用的业务类型 ID")
    common.add_argument("--business-type-name", default=None, help="客户/商机/报价单/合同导入使用的业务类型名称")
    common.add_argument("--common-id", default=None, help="线索池导入使用的 common_id")
    common.add_argument(
        "--business-type-map",
        default=None,
        help="批量时按模块指定业务类型，例如 customer=793,opportunity=商机业务类型2",
    )

    sub = parser.add_subparsers(dest="command", required=True)

    p = sub.add_parser("download-template", parents=[common])
    p.add_argument("--output", default=str(DEFAULT_OUT_DIR / "lead-template.xlsx"))
    p.set_defaults(func=download_template)

    p = sub.add_parser("generate-source-from-template")
    p.add_argument("--template", required=True)
    p.add_argument("--output", required=True)
    p.add_argument("--rows", type=int, default=2)
    p.add_argument("--rows-random", nargs=2, type=int, metavar=("MIN", "MAX"), help="随机生成条数区间")
    p.add_argument("--fill-mode", choices=["all", "minimal"], default="all")
    p.set_defaults(func=generate_source_from_template)

    p = sub.add_parser("import-flow", parents=[common])
    p.add_argument("--file", required=True)
    p.add_argument("--upload-name", default=None, help="上传到 OSS 时使用的文件名，默认使用本地文件名")
    p.add_argument("--faye-url", default=None)
    p.add_argument("--listen-faye", action=argparse.BooleanOptionalAction, default=True)
    p.add_argument("--validate-wait", type=int, default=45)
    p.add_argument("--import-wait", type=int, default=90)
    p.add_argument("--per-page", type=int, default=5)
    p.set_defaults(func=import_flow)

    p = sub.add_parser("full-flow", parents=[common])
    p.add_argument("--output-dir", default=str(DEFAULT_OUT_DIR))
    p.add_argument("--template-file", default=None, help="使用页面下载的原始导入模板作为生成基底")
    p.add_argument("--template-dir", default=None, help="页面下载原始导入模板目录，按表头自动匹配 CRM_*_导入模板.xlsx")
    p.add_argument("--rows", type=int, default=2)
    p.add_argument("--rows-random", nargs=2, type=int, metavar=("MIN", "MAX"), help="随机生成条数区间")
    p.add_argument("--fill-mode", choices=["all", "minimal"], default="all")
    p.add_argument("--upload-name", default=None, help="上传到 OSS 时使用的文件名，默认使用本地文件名")
    p.add_argument("--avoid-customer-ids", nargs="*", default=["8375665"], help="关联链随机选择时尽量避开的客户ID")
    p.add_argument("--faye-url", default=None)
    p.add_argument("--listen-faye", action=argparse.BooleanOptionalAction, default=True)
    p.add_argument("--validate-wait", type=int, default=45)
    p.add_argument("--import-wait", type=int, default=90)
    p.add_argument("--per-page", type=int, default=5)
    p.set_defaults(func=full_flow)

    p = sub.add_parser("batch-full-flow", parents=[common])
    p.add_argument(
        "--modules",
        default="all",
        help="逗号分隔模块，默认 all；例如 customer,contact,opportunity",
    )
    p.add_argument("--output-dir", default=str(DEFAULT_OUT_DIR / "batch"))
    p.add_argument("--template-file", default=None, help="使用页面下载的原始导入模板作为生成基底")
    p.add_argument("--template-dir", default=None, help="页面下载原始导入模板目录，按表头自动匹配 CRM_*_导入模板.xlsx")
    p.add_argument("--rows", type=int, default=1)
    p.add_argument("--rows-random", nargs=2, type=int, metavar=("MIN", "MAX"), help="随机生成条数区间")
    p.add_argument("--fill-mode", choices=["all", "minimal"], default="all")
    p.add_argument("--upload-name", default=None, help="上传到 OSS 时使用的文件名，默认使用本地文件名")
    p.add_argument("--avoid-customer-ids", nargs="*", default=["8375665"], help="关联链随机选择时尽量避开的客户ID")
    p.add_argument("--faye-url", default=None)
    p.add_argument("--listen-faye", action=argparse.BooleanOptionalAction, default=True)
    p.add_argument("--validate-wait", type=int, default=45)
    p.add_argument("--import-wait", type=int, default=90)
    p.add_argument("--per-page", type=int, default=3)
    p.add_argument("--continue-on-error", action=argparse.BooleanOptionalAction, default=True)
    p.set_defaults(func=batch_full_flow)

    p = sub.add_parser("error-template-flow", parents=[common])
    p.add_argument("--output-dir", default=str(DEFAULT_OUT_DIR / "error-template"))
    p.add_argument("--template-file", default=None, help="使用页面下载的原始导入模板作为生成基底")
    p.add_argument("--template-dir", default=None, help="页面下载原始导入模板目录，按表头自动匹配 CRM_*_导入模板.xlsx")
    p.add_argument("--rows", type=int, default=1)
    p.add_argument("--fill-mode", choices=["all", "minimal"], default="all")
    p.add_argument("--upload-name", default=None, help="上传到 OSS 时使用的文件名，默认使用本地文件名")
    p.add_argument("--faye-url", default=None)
    p.add_argument("--listen-faye", action=argparse.BooleanOptionalAction, default=True)
    p.add_argument("--validate-wait", type=int, default=45)
    p.add_argument("--import-wait", type=int, default=90)
    p.add_argument("--per-page", type=int, default=5)
    p.add_argument("--verify-error-zip", action="store_true", help="验证首次失败后的错误报告为 ZIP 后停止")
    p.set_defaults(func=error_template_flow)

    p = sub.add_parser("lead-error-template-flow", parents=[common])
    p.add_argument("--output-dir", default=str(DEFAULT_OUT_DIR / "error-template"))
    p.add_argument("--template-file", default=None, help="使用页面下载的原始导入模板作为生成基底")
    p.add_argument("--template-dir", default=None, help="页面下载原始导入模板目录，按表头自动匹配 CRM_*_导入模板.xlsx")
    p.add_argument("--rows", type=int, default=1)
    p.add_argument("--fill-mode", choices=["all", "minimal"], default="all")
    p.add_argument("--upload-name", default=None, help="上传到 OSS 时使用的文件名，默认使用本地文件名")
    p.add_argument("--faye-url", default=None)
    p.add_argument("--listen-faye", action=argparse.BooleanOptionalAction, default=True)
    p.add_argument("--validate-wait", type=int, default=45)
    p.add_argument("--import-wait", type=int, default=90)
    p.add_argument("--per-page", type=int, default=5)
    p.set_defaults(func=lead_error_template_flow)

    p = sub.add_parser("business-types", parents=[common])
    p.set_defaults(func=print_business_types)

    p = sub.add_parser("history", parents=[common])
    p.add_argument("--per-page", type=int, default=5)
    p.set_defaults(func=query_history)
    return parser


def main():
    args = build_parser().parse_args()
    resolve_config(args)
    args.func(args)


if __name__ == "__main__":
    main()
